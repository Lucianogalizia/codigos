#!/usr/bin/env python
# coding: utf-8

# In[2]:


#NO HACE FALTA EJECUTARLA-- ES PARA ELIMINAR LA INTERFAZ DONDE SE VE EL CODIGO------

%%javascript
require(["base/js/namespace"], function(Jupyter){
    // Construye y agrega un botón en la toolbar que oculte/mostrará las celdas de código
    Jupyter.toolbar.add_buttons_group([{
        'label'   : 'Mostrar/Ocultar código',
        'icon'    : 'fa-eye-slash',  // icono FontAwesome opcional
        'callback': function(){
            Jupyter.notebook.get_cells().forEach(function(cell){
                if(cell.cell_type === 'code'){
                    cell.element.find("div.input").toggle();
                }
            });
        }
    }]);
});


# In[1]:


import pandas as pd
from IPython.display import display
import numpy as np

# ===== Celda 1: Carga y visualización inicial del datasheet =====

# 1) Ruta a tu Excel (ajusta al path real)
ruta = r"C:\Users\ry16123\OneDrive - YPF\Escritorio\power BI\GUADAL- POWER BI\Inteligencia Artificial\Carga Programas de Pu\DATA SHEET\hoja_datos_LP-1682.xlsm"

# 2) Cargo la hoja 'Data Sheet' sin cabecera
df0 = pd.read_excel(ruta, sheet_name="Data Sheet", header=None, engine="openpyxl")
df  = df0.copy()

# 3) Defino posiciones de metadatos y los extraigo
meta_pos = {
    "POZO": (2,4), "BATERIA": (4,4), "EQUIPO": (6,4),
    "NETA ASOCIADA": (8,4), "DEFINICIÓN": (10,4), "MANIOBRAS (MOTIVO)": (12,4),
    "ANTECEDENTE_1": (17,4), "ANTECEDENTE_2": (18,4),
    "ANTECEDENTE_3": (19,4), "ANTECEDENTE_4": (20,4),
    "REQ_ESP_1": (22,4), "REQ_ESP_2": (23,4),
    "REQ_ESP_3": (24,4), "REQ_ESP_4": (25,4),
}
meta = {k: df0.iat[r,c] for k,(r,c) in meta_pos.items()}

# 4) Imprimo metadatos
print("=== METADATOS ===")
for k, v in meta.items():
    print(f"{k:25s}: {v}")

# 5) Extraigo las tablas principales
tubing_act = df.iloc[31:53, 3:8].copy()
tubing_act.columns = ["ELEMENTO","DIÁMETRO","PROFUNDIDAD","CANTIDAD","COMENTARIO"]
tubing_act = tubing_act.dropna(subset=["ELEMENTO"])

# ————— Primero convierto/trunco DIÁMETRO a 3 decimales —————
tubing_act['DIÁMETRO'] = pd.to_numeric(tubing_act['DIÁMETRO'], errors='coerce')
tubing_act['DIÁMETRO'] = np.trunc(tubing_act['DIÁMETRO'] * 1000) / 1000

# ————— Agrupo consecutivo las mismas filas de TUBING —————
# marco dónde cambian las 3 columnas clave
cambios = (
    (tubing_act["ELEMENTO"]    != tubing_act["ELEMENTO"].shift()) |
    (tubing_act["DIÁMETRO"]    != tubing_act["DIÁMETRO"].shift()) |
    (tubing_act["COMENTARIO"]  != tubing_act["COMENTARIO"].shift())
)
# asigno un ID de grupo creciente
tubing_act["_grp"] = cambios.cumsum()

# ahora agrego por grupo:
tubing_act = (
    tubing_act
    .groupby("_grp", as_index=False)
    .agg({
        "ELEMENTO":   "first",
        "DIÁMETRO":   "first",
        # PROFUNDIDAD se vacía cuando agrupamos más de 1 fila
        "PROFUNDIDAD": lambda s: "" if len(s)>1 else s.iloc[0],
        "CANTIDAD":   "sum",
        "COMENTARIO": "first"
    })
    .drop(columns="_grp")
)
#------------------------------
tubing_fin = df.iloc[31:54, 10:17].copy()
tubing_fin.columns = ["ELEMENTO","CONDICIÓN","DIÁMETRO","PROFUNDIDAD","CANTIDAD","COMENTARIO","LONGITUD ELEMENTO"]
tubing_fin = tubing_fin.dropna(subset=["ELEMENTO"])
tubing_fin['DIÁMETRO'] = pd.to_numeric(tubing_fin['DIÁMETRO'], errors='coerce')
tubing_fin['DIÁMETRO'] = np.trunc(tubing_fin['DIÁMETRO'] * 1000) / 1000  # <- truncado a 3 decimales

varillas_act = df.iloc[55:78, 3:8].copy()
varillas_act.columns = ["ELEMENTO","DIÁMETRO","PROFUNDIDAD","CANTIDAD","COMENTARIO"]
varillas_act = varillas_act.dropna(subset=["ELEMENTO"])
varillas_act["DIÁMETRO"] = pd.to_numeric(varillas_act["DIÁMETRO"], errors="coerce")
varillas_act["DIÁMETRO"] = varillas_act["DIÁMETRO"].map(lambda x: f"{x:.3f}")

# ————— Agrupo consecutivos LAS MISMAS FILAS varillas_act —————
# marco dónde cambian las 4 columnas clave
cambios_var = (
    (varillas_act["ELEMENTO"]   != varillas_act["ELEMENTO"].shift())   |
    (varillas_act["DIÁMETRO"]   != varillas_act["DIÁMETRO"].shift())   |
    (varillas_act["COMENTARIO"] != varillas_act["COMENTARIO"].shift())
)
# asigno un ID de grupo
varillas_act["_grp"] = cambios_var.cumsum()

# agrego por grupo: sumo cantidades, vacío profundidad si hay >1 fila
varillas_act = (
    varillas_act
    .groupby("_grp", as_index=False)
    .agg({
        "ELEMENTO":   "first",
        
        "DIÁMETRO":   "first",
        "PROFUNDIDAD": lambda s: "" if len(s)>1 else s.iloc[0],
        "CANTIDAD":   "sum",
        "COMENTARIO": "first"
    })
    .drop(columns="_grp")
)

varillas_fin = df.iloc[55:78, 10:19].copy()
varillas_fin.columns = ["ELEMENTO","CONDICIÓN","DIÁMETRO","PROFUNDIDAD","ACERO V/B","CUPLA SH/FS","ACERO CUPLA","CANTIDAD","COMENTARIO"]
varillas_fin = varillas_fin.dropna(subset=["ELEMENTO"])
varillas_fin["DIÁMETRO"] = pd.to_numeric(varillas_fin["DIÁMETRO"], errors="coerce")
varillas_fin["DIÁMETRO"] = varillas_fin["DIÁMETRO"].map(lambda x: f"{x:.3f}")


# 6) Vistazo rápido
print("\n=== Instalación Actual Tubing ===")
display(tubing_act.head())
print("=== Instalación Final Tubing ===")
display(tubing_fin.head())
print("=== Instalación Actual Varillas ===")
display(varillas_act.head())
print("=== Instalación Final Varillas ===")
display(varillas_fin.head())

#------AUTOCOMPLETADO DE CANTIDADES------


# 5bis) Rellenar por defecto CANTIDADES=1 en los elementos clave si estaban vacíos
keywords= ["SHEAR - OUT", "ZAPATO", "ANCLA", "NIPLE", "ZTO", "PACKER"]
# Construimos una expresión regular que busque cualquiera de las palabras
pattern = r"(?i)\b(?:" + "|".join(keywords) + r")\b|BBA\."

# Detectamos filas que coinciden y tienen CANTIDADES vacío o NaN
mask_vacios = (
    tubing_fin["ELEMENTO"].str.contains(pattern, regex=True, na=False)
    & (
        tubing_fin["CANTIDAD"].isna()
        | (tubing_fin["CANTIDAD"].astype(str).str.strip() == "")
    )
)

# Si hay alguna fila a corregir, las llenamos a 1 y avisamos
if mask_vacios.any():
    corregidos = tubing_fin.loc[mask_vacios, "ELEMENTO"].tolist()
    tubing_fin.loc[mask_vacios, "CANTIDAD"] = 1
    print("\n⚠️ ALERTA: Se asignó CANTIDADES=1 por defecto a los siguientes ELEMENTO(s) en Instalación Final Tubing:")
    for elem in corregidos:
        print(f"   • {elem}")

# ——— 5ter) Rellenar por defecto CANTIDADES=1 en Instalación Final Varillas ———

# Palabras clave a buscar en ELEMENTO
keywords_var = ["VASTAGO", "DISPOSITIVO ON-OFF"]
pattern_var  = r"(?i)(?:\b(?:%s)\b|BBA\.)" % "|".join(keywords_var)

# Filtramos filas donde ELEMENTO coincide y CANTIDAD está vacío o NaN
mask_var_vacios = (
    varillas_fin["ELEMENTO"].str.contains(pattern_var, regex=True, na=False)
    & (
        varillas_fin["CANTIDAD"].isna()
        | (varillas_fin["CANTIDAD"].astype(str).str.strip() == "")
    )
)

# Si hay filas a corregir, las llenamos con 1 y mostramos alerta
if mask_var_vacios.any():
    corregidos_var = varillas_fin.loc[mask_var_vacios, "ELEMENTO"].tolist()
    varillas_fin.loc[mask_var_vacios, "CANTIDAD"] = 1
    print("\n⚠️ ALERTA: Se asignó CANTIDADES=1 por defecto a los siguientes ELEMENTO(s) en Instalación Final Varillas:")
    for elem in corregidos_var:
        print(f"   • {elem}")


# —————— LÓGICAS DE ALERTA ——————

# Función auxiliar para chequear “CANTIDAD” vacía/nula
def alerta_cantidad(df_tabla, nombre_tabla):
    # Consideramos vacía tanto None/NaN como cadenas vacías o espacios
    faltan = df_tabla[
        df_tabla["CANTIDAD"].isna() |
        (df_tabla["CANTIDAD"].astype(str).str.strip() == "")
    ]["ELEMENTO"].tolist()
    if faltan:
        print(f"\n⚠️ ALERTA ({nombre_tabla}): hay elementos sin cantidad asignada:")
        for elem in faltan:
            print(f"   • ELEMENTO «{elem}» no tiene CANTIDAD")
    else:
        print(f"\n✅ ({nombre_tabla}): todas las filas tienen 'CANTIDAD' asignada.")

# 1) Alerta para Instalación Final de Tubing
alerta_cantidad(tubing_fin, "INSTALACIÓN FINAL TUBING")

# 2) Alerta para Instalación Final de Varillas
alerta_cantidad(varillas_fin, "INSTALACIÓN FINAL VARILLAS")

# ===== GENERACION DE PROGRAMAS DE PULLING. =====


import ipywidgets as widgets
from IPython.display import display, clear_output


# 1) Creamos dos botones y dos áreas de salida, pero mostramos solo el primero:
btn1   = widgets.Button(description="▶ Ejecutar Programa", button_style="primary")
out1   = widgets.Output()
btn2   = widgets.Button(description="▶ Predecir Tiempos", button_style="success")
out2   = widgets.Output()


# Ocultamos inicialmente el segundo
btn2.layout.display = 'none'
out2.layout.display = 'none'

# 2) Al hacer clic, ejecuto TODO el bloque “oculto”
def on_btn1_click(b):
    global varillas_act, varillas_fin, tubing_act, tubing_fin
    with out1:
        clear_output()   # limpia salidas previas

        import pandas as pd
        import ipywidgets as widgets
        

        # (Asume aquí que ya tienes df0, df, tubing_act, tubing_fin, varillas_act, varillas_fin y meta_pos de la Celda 1)

        # 1) Limpieza de varillas_act
        varillas_act = varillas_act[varillas_act["ELEMENTO"] != "ELEMENTO"]
        varillas_act["CANTIDAD"] = pd.to_numeric(varillas_act["CANTIDAD"], errors="coerce")
        varillas_act = varillas_act.dropna(subset=["CANTIDAD"])

        varillas_fin = varillas_fin[varillas_fin["ELEMENTO"] != "ELEMENTO"]
        varillas_fin["CANTIDAD"] = pd.to_numeric(varillas_fin["CANTIDAD"], errors="coerce")
        varillas_fin = varillas_fin.dropna(subset=["CANTIDAD"])

        # 2) Limpieza de tubing_act / tubing_fin
        tubing_act = tubing_act[tubing_act["ELEMENTO"] != "ELEMENTO"]
        tubing_act["CANTIDAD"] = pd.to_numeric(tubing_act["CANTIDAD"], errors="coerce")
        tubing_act = tubing_act.dropna(subset=["CANTIDAD"])

        tubing_fin = tubing_fin[tubing_fin["ELEMENTO"] != "ELEMENTO"]
        tubing_fin["CANTIDAD"] = pd.to_numeric(tubing_fin["CANTIDAD"], errors="coerce")
        tubing_fin = tubing_fin.dropna(subset=["CANTIDAD"])

        #import ipywidgets as widgets
        #from IPython.display import display, clear_output

        # ——————————————————————————————————————————————————————————————————————————————————
        # Dropdown de “Sistema” sin valor inicial
        # ——————————————————————————————————————————————————————————————————————————————————

        system_dropdown = widgets.Dropdown(
            options=[
                ("— Seleccione sistema —", None),
                ("PU - Pulling", "PU - Pulling"),
                ("FB - Flush by", "FB - Flush by"),
            ],
            value=None,
            description='Sistema:',
            style={'description_width': 'initial'}
        )

        program_dropdown = widgets.Dropdown(
            options=[],
            description='Programa:',
            style={'description_width': 'initial'}
        )

        subprog_dropdown = widgets.Dropdown(
            options=[],
            description='Sub‐programa:',
            style={'description_width': 'initial'}
        )

        out_ui = widgets.Output()

        def on_system_change(change):
            with out_ui:
                clear_output()
                # Al cambiar sistema, reseteamos los dropdowns de programa / subprograma:
                program_dropdown.options = []
                subprog_dropdown.options = []

                if change['new'] == 'PU - Pulling':
                    program_dropdown.options = [
                        'Cambio de BBA Ins + Tubing',
                        'Cambio BES X BM',
                        'Pesca y cambio bba',
                        'Cambio de bba sin mov tbg',
                        'Cambio bba + tbg + bayler',
                        'Pesca + cambio bba + tubilares',
                        'Recupero material',
                        'Cambio PCP X BM',
                        'Instalar BM',
                        'Parcial varrillas en pozo + saca desenrosque + tbg'
                    ]
                    display(program_dropdown)

                elif change['new'] == 'FB - Flush by':
                    program_dropdown.options = [
                        'Ajuste de medida',
                        'Cambio de BBA insertable',
                        'Cambio de bomba TH',
                        'Pesca y Enganche',
                        'Instalar Bomba BM'
                    ]
                    display(program_dropdown)

                else:
                    # Si dejó el placeholder (“— Seleccione sistema —”), no hacemos nada
                    pass

        # Conectamos el observador sólo una vez
        system_dropdown.observe(on_system_change, names='value')

        # ——————————————————————————————————————————————————————————————————————————————————
        # Cuando el usuario selecciona un “Programa”
        # ——————————————————————————————————————————————————————————————————————————————————

        def on_program_change(change):
            with out_ui:
                clear_output()
                # Mostrar siempre el dropdown de “Programa” para poder cambiarlo
                display(program_dropdown)
                subprog_dropdown.options = []

                if change['new'] == "Ajuste de medida":
                    generar_ajuste()
                elif change['new'] == "Cambio de BBA insertable":
                    generar_cambio_bba()
                elif change['new'] == "Cambio de bomba TH":
                    generar_cambio_th()
                elif change['new'] == "Instalar Bomba BM":
                    generar_instalar_bomba()
                elif change['new'] == "Pesca y Enganche":
                    generar_pesca_enganche()
                elif change['new'] == "Cambio BES X BM":
                    generar_cambio_bes_x_bm()
                elif change['new'] == "Pesca y cambio bba":
                    generar_cambio_Pesca_y_enganche_x_bm()
                elif change['new'] == "Cambio de bba sin mov tbg":
                    generar_cambio_bm()
                elif change['new'] == "Cambio bba + tbg + bayler":
                    generar_bayler_bm()
                elif change['new'] == "Pesca + cambio bba + tubilares":
                    generar_Pesca_Cambio_bba_tub_bm()
                elif change['new'] == "Cambio PCP X BM":
                    generar_cambio_PCP_x_bm()
                elif change['new'] == "Parcial varrillas en pozo + saca desenrosque + tbg":
                    generar_Sacar_desenrosque_tub_bm()
                elif change['new'] == "Instalar BM":
                    generar_instalar_bm()
                elif change['new'] == "Recupero material":
                    generar_recupero_bm()
                elif change['new'] == "Cambio de BBA Ins + Tubing":
                    # Solo este programa tiene “subprogramas”
                    subprog_dropdown.options = [
                        'Con ancla',
                        'Sin ancla',
                        'Mov de varillas con PH intermedia',
                        'Probador asentado buscando pérdida'
                    ]
                    display(subprog_dropdown)

        program_dropdown.observe(on_program_change, names='value')

        # ——————————————————————————————————————————————————————————————————————————————————
        # Cuando el usuario selecciona un “Sub‐programa”
        # ——————————————————————————————————————————————————————————————————————————————————

        def on_subprog_change(change):
            with out_ui:
                clear_output()
                # Volver a mostrar ambos dropdowns para que el usuario pueda cambiar selección
                display(program_dropdown, subprog_dropdown)

                if change['new'] == 'Con ancla':
                    generar_cambio_bba_tubing_con_ancla()
                elif change['new'] == 'Sin ancla':
                    generar_cambio_bba_tubing_sin_ancla()
                elif change['new'] == 'Mov de varillas con PH intermedia':
                    generar_mov_varillas_ph_intermedia()
                elif change['new'] == 'Probador asentado buscando pérdida':
                    generar_probador_asentado()

        subprog_dropdown.observe(on_subprog_change, names='value')

        # ——————————————————————————————————————————————————————————————————————————————————
        # GENERACION DE PROGRAMAS DE FB
        # ——————————————————————————————————————————————————————————————————————————————————

        display(system_dropdown, out_ui)

        # 3) Función para “Ajuste de medida”
        def generar_ajuste():
            meta = {k: df0.iat[r,c] for k,(r,c) in meta_pos.items()}
            m = []
            # — EQUIPO EN TRANSPORTE
            desc = f"Transportar a {meta['POZO']}. "
            ants = [meta[f"ANTECEDENTE_{i}"] for i in range(1,5) if pd.notna(meta[f"ANTECEDENTE_{i}"])]
            if ants:
                desc += "Tener en cuenta los antecedentes del pozo: " + ", ".join(ants) + ". "
            desc += f"La definición actual del pozo es: {meta['DEFINICIÓN']}. La maniobra a realizar es {meta['MANIOBRAS (MOTIVO)']}. "
            reqs = [meta[f"REQ_ESP_{i}"] for i in range(1,5) if pd.notna(meta[f"REQ_ESP_{i}"])]
            if reqs:
                desc += "Considerar los siguientes requerimientos: " + "; ".join(reqs) + "."
            m.append({
                "MANIOBRA NORMALIZADA":"EQUIPO EN TRANSPORTE","PUNTO PROGRAMA":1,
                "DESCRIPCIÓN":desc,"ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP10","ACTIVITY_SUBCODE":"200","TIEMPO":""
            })
            # — MONTAJE EQUIPO
            m.append({
                "MANIOBRA NORMALIZADA":"MONTAJE EQUIPO","PUNTO PROGRAMA":2,
                "DESCRIPCIÓN":"Verificar presiones por directa y por entrecaño. Desarmar puente de producción. Montar equipo según procedimiento.",
                "ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP10","ACTIVITY_SUBCODE":"201","TIEMPO":""
            })

            
            # — ACONDICIONA PARA SACAR VARILLAS
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","PUNTO PROGRAMA":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. Tomar peso de sarta, y registrar en OW.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
            # — SACA VARILLAS (ajuste de medida)
            parts=[]
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                comm = row["COMENTARIO"] if isinstance(row["COMENTARIO"], str) else ""
                up = comm.upper()
                mode = "(simple)" if "SIMPLE" in up else "(doble)" if "DOBLE" in up else ""
                parts.append(f"{qty} {elem} {mode}".strip())
            design = " + ".join(parts)
            desc_sv = (
                "Retirar vástago completo SIN DESCLAVAR BOMBA, reemplazar por nuevo. "
                "AGREGAR UN trozo DE 2 FT para realizar ajuste de medida. "
                f"Diseño en el pozo: {design}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído."
            )
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS","PUNTO PROGRAMA":5,
                "DESCRIPCIÓN":desc_sv,"ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""
            })
            # — ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH","PUNTO PROGRAMA":6,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })
            # — PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD","PUNTO PROGRAMA":7,
                "DESCRIPCIÓN":"Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })
            # — VARIOS
            m.append({
                "MANIOBRA NORMALIZADA":"VARIOS","PUNTO PROGRAMA":8,
                "DESCRIPCIÓN":"Tareas varias.","ACTIVITY_PHASE":"SPV","ACTIVITY_CODE":"SPV","ACTIVITY_SUBCODE":"SPV","TIEMPO":""
            })
            
            # — DESMONTA EQUIPO
            m.append({
                "MANIOBRA NORMALIZADA":"DESMONTA EQUIPO","PUNTO PROGRAMA":10,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado.",
                "ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP11","ACTIVITY_SUBCODE":"202","TIEMPO":""
            })

            df_prog = pd.DataFrame(m)
            display(df_prog)
            df_prog.to_excel("ajuste_de_medida_program.xlsx", index=False)
            print("✅ Guardado: ajuste_de_medida_program.xlsx")


        # 4) Función para “Cambio de BBA insertable”
        def generar_cambio_bba():
            meta = {k: df0.iat[r,c] for k,(r,c) in meta_pos.items()}
            m = []
            # — 1) EQUIPO EN TRANSPORTE
            desc = f"Transportar a {meta['POZO']}. "
            ants = [meta[f"ANTECEDENTE_{i}"] for i in range(1,5) if pd.notna(meta[f"ANTECEDENTE_{i}"])]
            if ants:
                desc += "Tener en cuenta los antecedentes del pozo: " + ", ".join(ants) + ". "
            desc += f"La definición actual del pozo es: {meta['DEFINICIÓN']}. La maniobra a realizar es {meta['MANIOBRAS (MOTIVO)']}. "
            reqs = [meta[f"REQ_ESP_{i}"] for i in range(1,5) if pd.notna(meta[f"REQ_ESP_{i}"])]
            if reqs:
                desc += "Considerar los siguientes requerimientos: " + "; ".join(reqs) + "."
            m.append({
                "MANIOBRA NORMALIZADA":"EQUIPO EN TRANSPORTE","PUNTO PROGRAMA":1,
                "DESCRIPCIÓN":desc,"ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP10","ACTIVITY_SUBCODE":"200","TIEMPO":""
            })
            # — 2) MONTAJE EQUIPO
            m.append({
                "MANIOBRA NORMALIZADA":"MONTAJE EQUIPO","PUNTO PROGRAMA":2,
                "DESCRIPCIÓN":"Verificar presiones por directa y por entrecaño. Desarmar puente de producción. Montar equipo según procedimiento.",
                "ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP10","ACTIVITY_SUBCODE":"201","TIEMPO":""
            })

            
            # — 6) ACONDICIONA PARA SACAR VARILLAS (con PH inicial)
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","PUNTO PROGRAMA":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. Tomar peso de sarta, y registrar en OW. Realizar PH inicial con 1000 psi e informar resultados, de ser positiva continuar con maniobras.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
            # — 7) SACA VARILLAS (solo peso/anomalías)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS","PUNTO PROGRAMA":7,
                "DESCRIPCIÓN":"Tareas varias durante saca de varillas",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""
            })

            # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                parts.append(f"{qty} {elem} {dia}in {modo}".strip())
            diseño = " + ".join(parts)

            desc_sv_simple = (
                "Desclavar bomba."
                "Sacar sarta en tiro simple, desarmando, completando pozo. Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído."
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "PUNTO PROGRAMA":        10,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })

            # —12) ACONDICIONAMIENTO PARA BAJAR VARILLAS
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA BAJAR VARILLAS","PUNTO PROGRAMA":12,
                "DESCRIPCIÓN":"Acondicionar boca de pozo + herramientas de v/b.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"254","TIEMPO":""
            })
            # —13) BAJA VARILLAS
            m.append({
                "MANIOBRA NORMALIZADA":"BAJA VARILLAS","PUNTO PROGRAMA":13,
                "DESCRIPCIÓN":"Tareas varias durante bajado de varillas.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"255","TIEMPO":""
            })
            # —14) BAJA VARILLAS EN SIMPLE (punto 14)
            parts = []
            # iteramos varillas_fin en orden inverso
            for _, row in varillas_fin.iloc[::-1].iterrows():
                # cantidad
                qty = int(row["CANTIDAD"])  
                # elemento y condición
                elem = row["ELEMENTO"]
                cond = row["CONDICIÓN"] if pd.notnull(row["CONDICIÓN"]) else ""
                # profundidad en mts
                depth = f" {int(row['PROFUNDIDAD'])} mts" if pd.notnull(row["PROFUNDIDAD"]) else ""
                # diámetro
                diam = f" de {row['DIÁMETRO']}" if pd.notnull(row["DIÁMETRO"]) else ""
                # acero V/B
                acero_vb = f", {row['ACERO V/B']}" if pd.notnull(row["ACERO V/B"]) else ""
                # cupla SH/FS
                cupla = f", cupla {row['CUPLA SH/FS']}" if pd.notnull(row["CUPLA SH/FS"]) else ""
                # acero cupla
                acero_cupla = f", {row['ACERO CUPLA']}" if pd.notnull(row["ACERO CUPLA"]) else ""
                # comentario (solo para marcar BAJA EN SIMPLE)
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                comm_tag = " (BAJA EN SIMPLE)" if "BAJA EN SIMPLE" in comm.upper() else ""

                # armamos la pieza
                piece = f"{qty} {elem}"
                if cond:
                    piece += f" {cond}"
                piece += depth + diam + acero_vb + cupla + acero_cupla + comm_tag

                parts.append(piece.strip())

            diseño = " + ".join(parts)

            m.append({
                "MANIOBRA NORMALIZADA":   "BAJA VARILLAS EN SIMPLE",
                "PUNTO PROGRAMA":         14,
                "DESCRIPCIÓN": (
                    "Tomar datos de bomba y bajarla + sarta de v/b limpiando todas las conexiones "
                    "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                    "de acero de varillas. Diseño a bajar: " + diseño + "."
                ),
                "ACTIVITY_PHASE": "SP04",
                "ACTIVITY_CODE":  "SP16",
                "ACTIVITY_SUBCODE":"255",
                "TIEMPO": ""
            })
            
            
            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "PUNTO PROGRAMA":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "PUNTO PROGRAMA":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })
            
            # —15) VARIOS
            m.append({
                "MANIOBRA NORMALIZADA":"VARIOS","PUNTO PROGRAMA":15,
                "DESCRIPCIÓN":"Tareas varias.","ACTIVITY_PHASE":"SPV","ACTIVITY_CODE":"SPV","ACTIVITY_SUBCODE":"SPV","TIEMPO":""
            })
            
            # —17) DESMONTA EQUIPO
            m.append({
                "MANIOBRA NORMALIZADA":"DESMONTA EQUIPO","PUNTO PROGRAMA":17,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, material sobrante y locación. Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado.",
                "ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP11","ACTIVITY_SUBCODE":"202","TIEMPO":""
            })

            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("cambio_bba_insertable_program.xlsx", index=False)
            print("✅ Guardado: cambio_bba_insertable_program.xlsx")

        def generar_cambio_th():
            # copiamos la lógica completa de generar_cambio_bba(), 
            # pero cambiamos sólo la descripción de los pts 10 y 14:
            meta = {k: df0.iat[r,c] for k,(r,c) in meta_pos.items()}
            m = []
            # — 1) EQUIPO EN TRANSPORTE
            desc = f"Transportar a {meta['POZO']}. "
            ants = [meta[f"ANTECEDENTE_{i}"] for i in range(1,5) if pd.notna(meta[f"ANTECEDENTE_{i}"])]
            if ants:
                desc += "Tener en cuenta los antecedentes del pozo: " + ", ".join(ants) + ". "
            desc += f"La definición actual del pozo es: {meta['DEFINICIÓN']}. La maniobra a realizar es {meta['MANIOBRAS (MOTIVO)']}. "
            reqs = [meta[f"REQ_ESP_{i}"] for i in range(1,5) if pd.notna(meta[f"REQ_ESP_{i}"])]
            if reqs:
                desc += "Considerar los siguientes requerimientos: " + "; ".join(reqs) + "."
            m.append({
                "MANIOBRA NORMALIZADA":"EQUIPO EN TRANSPORTE","PUNTO PROGRAMA":1,
                "DESCRIPCIÓN":desc,"ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP10","ACTIVITY_SUBCODE":"200","TIEMPO":""
            })
            # — 2) MONTAJE EQUIPO
            m.append({
                "MANIOBRA NORMALIZADA":"MONTAJE EQUIPO","PUNTO PROGRAMA":2,
                "DESCRIPCIÓN":"Verificar presiones por directa y por entrecaño. Desarmar puente de producción. Montar equipo según procedimiento.",
                "ACTIVITY_PHASE":"S01","ACTIVITY_CODE":"SP10","ACTIVITY_SUBCODE":"201","TIEMPO":""
            })

            
            # — 6) ACONDICIONA PARA SACAR VARILLAS (con PH inicial)
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","PUNTO PROGRAMA":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. Tomar peso de sarta, y registrar en OW. Realizar PH inicial con 1000 psi e informar resultados, de ser positiva continuar con maniobras.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
            # — 7) SACA VARILLAS (solo peso/anomalías)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS","PUNTO PROGRAMA":7,
                "DESCRIPCIÓN":"Tareas varias durante sacado de varillas.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""
            })
            
            # —10) SACA VARILLAS EN SIMPLE con texto TH
            parts = []
            for _, row in varillas_act.iterrows():
                qty, elem, dia = int(row["CANTIDAD"]), row["ELEMENTO"], row["DIÁMETRO"]
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                mode = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                parts.append(f"{qty} {elem} {dia}in {mode}".strip())
            diseño_act = " + ".join(parts)
            desc_th_sv = (
                "Libra ON-OFF PISTON."
                "Sacar sarta, desarmando, completando pozo. Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño_act}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído."
            )
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS EN SIMPLE",
                "PUNTO PROGRAMA":10,
                "DESCRIPCIÓN":desc_th_sv,
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""
            })


            # — 8) CIRCULA
            m.append({
                "MANIOBRA NORMALIZADA":"CIRCULA","PUNTO PROGRAMA":8,
                "DESCRIPCIÓN":"Circular pozo, 1.5 veces la capacidad de tubing por directa hasta retorno limpio para asegurar limpieza de los materiales extraídos. Si no se observa circulación informar si es por punta de instalación obstruida o porque el pozo admite.",
                "ACTIVITY_PHASE":"SP05","ACTIVITY_CODE":"SP20","ACTIVITY_SUBCODE":"218","TIEMPO":""
            })
            
            # —12) ACONDICIONAMIENTO PARA BAJAR VARILLAS
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA BAJAR VARILLAS","PUNTO PROGRAMA":12,
                "DESCRIPCIÓN":"Acondicionar boca de pozo + herramientas de v/b.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"254","TIEMPO":""
            })
            # —13) BAJA VARILLAS
            m.append({
                "MANIOBRA NORMALIZADA":"BAJA VARILLAS","PUNTO PROGRAMA":13,
                "DESCRIPCIÓN":"Tareas varias durante bajado de varillas.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"255","TIEMPO":""
            })

            # —14) BAJA VARILLAS EN SIMPLE con texto TH
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty = int(row["CANTIDAD"]); elem = row["ELEMENTO"]
                condicion = row.get("CONDICIÓN","") or ""
                depth = f" {int(row['PROFUNDIDAD'])} mts" if pd.notnull(row["PROFUNDIDAD"]) else ""
                diam   = f" de {row['DIÁMETRO']}" if pd.notnull(row["DIÁMETRO"]) else ""
                comm = row["COMENTARIO"] if pd.notna(row["COMENTARIO"]) else ""
                mode = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                parts.append(f"{qty} {elem}{(' '+condicion) if condicion else ''}{depth}{diam}{mode}".strip())
            diseño_fin = " + ".join(parts)
            desc_th_bv = (
                "Tomar datos de bomba TH y bajarla + sarta de v/b en doble limpiando todas las conexiones con detergente biodegradable. "
                "Realizar control de torque cada 15 varillas según grado de acero de varillas. "
                f"Diseño a bajar: {diseño_fin}."
            )
            m.append({
                "MANIOBRA NORMALIZADA":"BAJA VARILLAS EN SIMPLE",
                "PUNTO PROGRAMA":14,
                "DESCRIPCIÓN":desc_th_bv,
                "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"255","TIEMPO":""
            })
            
            
            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "PUNTO PROGRAMA":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "PUNTO PROGRAMA":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })
            
            # —15) VARIOS
            m.append({
                "MANIOBRA NORMALIZADA":"VARIOS","PUNTO PROGRAMA":15,
                "DESCRIPCIÓN":"Tareas varias.","ACTIVITY_PHASE":"SPV","ACTIVITY_CODE":"SPV","ACTIVITY_SUBCODE":"SPV","TIEMPO":""
            })
            
            # —17) DESMONTA EQUIPO
            m.append({
                "MANIOBRA NORMALIZADA":"DESMONTA EQUIPO","PUNTO PROGRAMA":17,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado.",
                "ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP11","ACTIVITY_SUBCODE":"202","TIEMPO":""
            })

            df_th = pd.DataFrame(m)
            display(df_th)
            df_th.to_excel("cambio_bomba_th_program.xlsx", index=False)
            print("✅ Guardado: cambio_bomba_th_program.xlsx")

        def generar_instalar_bomba():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

                     
            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo}. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

                 
            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo, montar piso de trabajo + herramientas de v/b."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })

            # Maniobra 14: BAJA VARILLAS
            desc14 = "Tareas varias durante bajada de varillas."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       14,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc14
            })

           # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de bomba y bajarla + sarta de v/b en simple limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. Realizar medida correctamente "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       15,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc15
            })


            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "Punto programa":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "Punto programa":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi."
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })



            # Maniobra 16: VARIOS
            desc16 = "Tareas varias."
            m.append({
                "MANIOBRA NORMALIZADA": "VARIOS",
                "ACTIVITY_PHASE":       "SPV",
                "ACTIVITY_CODE":        "SPV",
                "ACTIVITY_SUBCODE":     "SPV",
                "Punto programa":       16,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc16
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("InstalarBM_FB_program.xlsx", index=False)
        
        
        def generar_pesca_enganche():
            meta = {k: df0.iat[r,c] for k,(r,c) in meta_pos.items()}
            m = []
            # 1–5: idénticos a otros programas
            m.append({ "MANIOBRA NORMALIZADA":"EQUIPO EN TRANSPORTE","PUNTO PROGRAMA":1,
                       "DESCRIPCIÓN":(
                           f"Transportar a {meta['POZO']}. "
                           + ("Tener en cuenta los antecedentes del pozo: " +
                              ", ".join(meta[f"ANTECEDENTE_{i}"] for i in range(1,5)
                                        if pd.notna(meta[f"ANTECEDENTE_{i}"])) + ". ")
                           + f"La definición actual del pozo es: {meta['DEFINICIÓN']}. "
                           + f"La maniobra a realizar es {meta['MANIOBRAS (MOTIVO)']}. "
                           + ("Considerar los siguientes requerimientos: " +
                              "; ".join(meta[f"REQ_ESP_{i}"] for i in range(1,5)
                                        if pd.notna(meta[f"REQ_ESP_{i}"])) + ".")
                       ),
                       "ACTIVITY_PHASE":"S01","ACTIVITY_CODE":"SP10","ACTIVITY_SUBCODE":"200","TIEMPO":"" })
            for nm,phase,code,sub,pt,desc in [
                ("MONTAJE EQUIPO","SP01","SP10","201",2,
                   "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. Montar equipo según procedimiento."),
                ("ACONDICIONA PARA SACAR VARILLAS","SP03","SP24","250",4,
                   "Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. "
                   "Tomar peso de sarta, y registrar en OW. Realizar PH inicial con 1000 psi e informar resultados, "
                   "de ser positiva continuar con maniobras."),
                ("SACA VARILLAS","SP03","SP24","251",5,
                   "Tareas varias durante sacado de varillas."),

            ]:
                m.append({
                    "MANIOBRA NORMALIZADA":nm,
                    "PUNTO PROGRAMA":pt,
                    "DESCRIPCIÓN":desc,
                    "ACTIVITY_PHASE":phase,
                    "ACTIVITY_CODE":code,
                    "ACTIVITY_SUBCODE":sub,
                    "TIEMPO":""
                })
                    
            # —10) SACA VARILLAS EN PESCA
            parts_act = []
            for _,r in varillas_act.iterrows():
                qty, elem, dia = int(r["CANTIDAD"]), r["ELEMENTO"], r["DIÁMETRO"]
                parts_act.append(f"{qty} {elem} {dia}in")
            diseño_act = " + ".join(parts_act)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS EN PESCA",
                "PUNTO PROGRAMA":6,
                "DESCRIPCIÓN":(
                    "Saca varillas en simple hasta punto de pesca, completando pozo. Desarmar componentes que requieran reemplazo. "
                    f"Diseño a extraer: {diseño_act}. "
                    "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                    "Registrar evidencia fotográfica del estado del material y punto de pesca. Asentar en OW grado de acero del material extraído."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""
            })

            # —11) SACA VARILLAS EN PESCA EN DOBLE
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS EN PESCA EN SIMPLE",
                "PUNTO PROGRAMA":7,
                "DESCRIPCIÓN":(
                    "Bajar sarta de varillas en simple + pescador y Pescar. Desclavar bomba. "
                    "Sacar varillas pescadas en simple, desarmando conexión, completando pozo. "
                    "Desarmar componentes que requieran reemplazo. Reemplazar 5 varillas por arriba y por debajo en zona de pesca."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""
            })

            # — 8) CIRCULA
            m.append({
                "MANIOBRA NORMALIZADA":"CIRCULA","PUNTO PROGRAMA":8,
                "DESCRIPCIÓN":"Circular pozo, 1.5 veces la capacidad de tubing por directa hasta retorno limpio para asegurar limpieza de los materiales extraídos. Si no se observa circulación informar si es por punta de instalación obstruida o porque el pozo admite.",
                "ACTIVITY_PHASE":"SP05","ACTIVITY_CODE":"SP20","ACTIVITY_SUBCODE":"218","TIEMPO":""
            })

            # —15–17: acondiciona/baja varillas
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA BAJAR VARILLAS","PUNTO PROGRAMA":10,
                "DESCRIPCIÓN":"Acondicionar boca de pozo para comenzar a bajar varillas.",
                "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"254","TIEMPO":""})
            m.append({
                "MANIOBRA NORMALIZADA":"BAJA VARILLAS","PUNTO PROGRAMA":11,
                "DESCRIPCIÓN":"Tareas varias durante bajado de varillas.",
                "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"255","TIEMPO":""})
            # 17) BAJA VARILLAS EN SIMPLE
             # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notna(row["COMENTARIO"]) else ""
                comm_tag = " (BAJA EN SIMPLE)" if "BAJA EN SIMPLE" in comm.upper() else ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de bomba y bajarla + sarta de v/b en simple limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "PUNTO PROGRAMA":       12,
                "DESCRIPCIÓN":          desc15,
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "TIEMPO":               "",

            })

            # —18–22…
            for nm,phase,code,sub,pt,desc in [
                ("ACONDICIONA PARA PH","SP03","SP13","259",13,"Acondicionar superficie para realizar prueba hidráulica."),
                ("PRUEBA DE HERMETICIDAD","SP03","SP13","205",14,
                 "Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."),
                ("VARIOS","SPV","SPV","SPV",15,"Tareas varias."),
                ("DESMONTA EQUIPO","SP01","SP11","202",17,
                 "Acondicionar boca de pozo, material sobrante y locación. Desmontar. "
                 "Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                 "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado.")
            ]:
                m.append({
                    "MANIOBRA NORMALIZADA":nm,
                    "PUNTO PROGRAMA":pt,
                    "DESCRIPCIÓN":desc,
                    "ACTIVITY_PHASE":phase,
                    "ACTIVITY_CODE":code,
                    "ACTIVITY_SUBCODE":sub,
                    "TIEMPO":""
                })

            df_pesca = pd.DataFrame(m)
            display(df_pesca)
            df_pesca.to_excel("pesca_y_enganche_program.xlsx", index=False)
            print("✅ Guardado: pesca_y_enganche_program.xlsx")

        #---------------------- COMIENZA LOGICA PARA EQUIPOS DE PULLING--------------------------------
        
        import pandas as pd
        import math
        import ipywidgets as widgets
        
        WALL_THICKNESS = {
            2.875: 0.434,
            3.5:   0.508,
        }
        CONVERSION_PIES_METROS = 3.28084
        COEF_POISSON        = 0.3
        MODULO_YOUNG        = 30_000_000
        COEF_EXPANSION      = 0.0000069
        GRADIENTE_FLUIDO    = 0.5
        TEMP_SUPERFICIE_C   = 30
        TEMP_MEDIA_C        = 15
        NIVEL_ESTATICO_PIE  = 656.17

        def _build_base_maniobras(meta, varillas_act, tubing_act):
            """Añade maniobras 1–17 al listado m y lo devuelve."""
            m = []
            # 1) EQUIPO EN TRANSPORTE
            desc = f"Transportar a {meta['POZO']}. "
            ants = [meta[f"ANTECEDENTE_{i}"] for i in range(1,5) if pd.notna(meta[f"ANTECEDENTE_{i}"])]
            if ants:
                desc += "Tener en cuenta los antecedentes del pozo: " + ", ".join(ants) + ". "
            desc += f"La definición actual del pozo es: {meta['DEFINICIÓN']}. La maniobra a realizar es {meta['MANIOBRAS (MOTIVO)']}. "
            reqs = [meta[f"REQ_ESP_{i}"] for i in range(1,5) if pd.notna(meta[f"REQ_ESP_{i}"])]
            if reqs:
                desc += "Considerar los siguientes requerimientos: " + "; ".join(reqs) + "."
            m.append({
                "MANIOBRA NORMALIZADA":"EQUIPO EN TRANSPORTE","PUNTO PROGRAMA":1,
                "DESCRIPCIÓN":desc,"ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP10","ACTIVITY_SUBCODE":"200","TIEMPO":""
            })

            # 2) MONTAJE EQUIPO
            m.append({ "MANIOBRA NORMALIZADA":"MONTAJE EQUIPO","PUNTO PROGRAMA":2,
                "DESCRIPCIÓN":"Verificar presiones por directa y por entrecaño. "
                              "Desarmar puente de producción. Montar equipo según procedimiento.",
                "ACTIVITY_PHASE":"S01","ACTIVITY_CODE":"SP10","ACTIVITY_SUBCODE":"201","TIEMPO":"" })


            # 4) DESARMA BDP
            m.append({ "MANIOBRA NORMALIZADA":"DESARMA BDP","PUNTO PROGRAMA":4,
                "DESCRIPCIÓN":"Desarmar BDP.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP34","ACTIVITY_SUBCODE":"210","TIEMPO":"" })


            # 6) ACONDICIONA PARA SACAR VARILLAS
            m.append({ "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","PUNTO PROGRAMA":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. "
                              "Tomar peso de sarta, y registrar en OW. Realizar PH inicial con 1000 psi e informar resultados.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":"" })

            # 7) SACA VARILLAS
            m.append({ "MANIOBRA NORMALIZADA":"SACA VARILLAS","PUNTO PROGRAMA":7,
                "DESCRIPCIÓN":"Tareas varias durante sacado de varillas",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":"" })

            
            # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm  = row["COMENTARIO"] or ""
                # convierto a float para comparar
                try:
                    dia_val = float(dia)
                except:
                    dia_val = None

                # 1) lógica especial para PESO
                if "PESO" in elem.upper() and dia_val is not None:
                    if abs(dia_val - 1.5) < 1e-3:
                        modo = "(simple)"
                    elif abs(dia_val - 1.125) < 1e-3:
                        modo = "(doble)"
                    else:
                        # cae al modo por defecto
                        modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                else:
                    # 2) modo por comentario
                    modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""

                parts.append(f"{qty} {elem} {dia}in {modo}".strip())

            diseño = " + ".join(parts)

            desc_sv_simple = (
                "Desclavar bomba. "
                "Sacar sarta , desarmando, completando pozo. Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído."
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "PUNTO PROGRAMA":        8,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })

            # 11) ACONDICIONAMIENTO PARA SACAR CAÑOS
            m.append({ "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA SACAR CAÑOS","PUNTO PROGRAMA":11,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, completar con ASDF, desempaquetar y montar conjunto BOP.Montar piso de trabajo.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"252","TIEMPO":"" })

            # 12) SACA TUBING
            m.append({ "MANIOBRA NORMALIZADA":"SACA TUBING","PUNTO PROGRAMA":12,
                "DESCRIPCIÓN":"Tareas generales durante la sacada de tubing.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })

            # 13) SACA TUBING EN DOBLE with anchor
            parts_tub = [
                f"{int(r['CANTIDAD'])} {r['ELEMENTO']} {r['DIÁMETRO']}in "
                f"({'DOBLE' if 'DOBLE' in str(r['COMENTARIO']).upper() else 'SIMPLE'})"
                for _,r in tubing_act.iterrows()
            ]
            diseño_tub = " + ".join(parts_tub)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA TUBING EN DOBLE","PUNTO PROGRAMA":13,
                "DESCRIPCIÓN":(
                    "Librar ANCLA. Sacar columna buscando pérdida y completando pozo. Desarmar componentes que requieran reemplazo. "
                    f"Diseño a extraer: {diseño_tub}. Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído. Solicitar envío de bomba a taller de Pecom de inmediato para desarmar e inspeccionar. Indicar si evidencia falla visible. Asentar número de bomba y estado de cabezal y filtro.."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })



            # 15) ACONDICIONAMIENTO PARA BAJAR CAÑOS
            #m.append({ "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA BAJAR CAÑOS","PUNTO PROGRAMA":15,
             #   "DESCRIPCIÓN":"Completar pozo con ASDF y retirar BOP. Empaquetar pozo.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"256","TIEMPO":"" })

            # 16) BAJA TUBING
            m.append({ "MANIOBRA NORMALIZADA":"BAJA TUBING","PUNTO PROGRAMA":16,
                "DESCRIPCIÓN":"Tareas generales durante la bajada de tubing.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"257","TIEMPO":"" })

            #17)BAJA TUBING EN SIMPLE Construyo “INSTALACIÓN FINAL tubing” a partir de tubing_fin
            partes_fin_tub = []
            for _, r in tubing_fin.iterrows():
                # Sólo si hay un ELEMENTO válido
                if isinstance(r["ELEMENTO"], str) and r["ELEMENTO"].strip():
                    cant       = int(r["CANTIDAD"]) if pd.notna(r["CANTIDAD"]) else 0
                    elem       = r["ELEMENTO"]
                    condicion  = r["CONDICIÓN"] if pd.notna(r["CONDICIÓN"]) else ""
                    diam       = r["DIÁMETRO"]
                    profundidad = r["PROFUNDIDAD"]
                    comentario_libre = r["COMENTARIO"] if pd.notna(r["COMENTARIO"]) else ""
                    # Armo la parte básica: CANTIDAD + ELEMENTO + CONDICIÓN + DIÁMETRO
                    parte = f"{cant} {elem}"
                    if condicion:
                        parte += f" {condicion}"
                    parte += f" {diam}in"

                    # Solo agrego "en X m" si PROFUNDIDAD NO es NaN
                    if pd.notna(profundidad):
                        parte += f" en {int(profundidad)}m"
                    if comentario_libre:
                        parte += f" ({comentario_libre})"
                    partes_fin_tub.append(parte)

            # Invierto el orden
            partes_fin_tub.reverse()

            # Concateno con " + "
            diseño_fin_tub = " + ".join(partes_fin_tub) if partes_fin_tub else ""

            desc9 = (
                "Profundizar columna de TBG calibrando, midiendo, limpiando, engrasando y torqueando las conexiones. "
                "Bajar instalación. "
                f"Diseño a bajar: {diseño_fin_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar número de bomba bajado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING EN SIMPLE",
                "PUNTO PROGRAMA":       17,
                "DESCRIPCIÓN":          desc9,
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "TIEMPO":               "",

            })


        # 19) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH","PUNTO PROGRAMA":18,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })

            # 20) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD","PUNTO PROGRAMA":19,
                "DESCRIPCIÓN":"Realizar PH inicial, intermedia y final con 1000, 900 y 800 psi respectivamente. Registrar la misma en OpenWells.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })
            return m

        # ——————————————————————————————
        def _añadir_fija_y_finalizar(m, tubing_fin, prof_m):
            """Calcula y añade la maniobra 18 + las 19–28, luego muestra y guarda."""
            # — 18) FIJA ANCLA —
            ancla_row = tubing_fin[tubing_fin["ELEMENTO"].str.contains("ANCLA", case=False, na=False)]
            #diam_ext = float(ancla_row["DIÁMETRO"].iloc[0])
            diam_ext = 2.875
            thickness = WALL_THICKNESS.get(diam_ext, 0.434)
            diam_int  = diam_ext - thickness

            prof_ft    = prof_m * CONVERSION_PIES_METROS
            dynamic_m  = max(prof_m - 200, 0)
            dynamic_ft = dynamic_m * CONVERSION_PIES_METROS

            #area_ext      = math.pi/4 * diam_ext**2
            #area_int      = math.pi/4 * diam_int**2
            #seccion_pared = area_ext - area_int

            ratio = ((COEF_POISSON * dynamic_ft / prof_ft) + (1-2*COEF_POISSON)
                     ) if prof_ft>0 else (1-2*COEF_POISSON)
            F1 = 6.4918 * dynamic_ft * GRADIENTE_FLUIDO * ratio
            F2 = MODULO_YOUNG * COEF_EXPANSION * ((TEMP_SUPERFICIE_C - TEMP_MEDIA_C)/2) * 1.812
            F3 = 6.4918 * GRADIENTE_FLUIDO * (( 1.812/diam_ext**2)*NIVEL_ESTATICO_PIE)

            tension_total   = F1 + F2 - F3
            estiramiento_in = 0.22 * (prof_ft/1000)*(tension_total/1000)
            est_cm          = estiramiento_in * 2.54

            desc_fija = (
                f"Fijar ancla con {tension_total:.2f} lbs "
                f"y {est_cm:.2f} cm de estiramiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA":"FIJA ANCLA","PUNTO PROGRAMA":20,
                "DESCRIPCIÓN":desc_fija,
                "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP17","ACTIVITY_SUBCODE":"211","TIEMPO":""
            })



            # 21) ACONDICIONAMIENTO PARA BAJAR VARILLAS
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA BAJAR VARILLAS","PUNTO PROGRAMA":21,
                "DESCRIPCIÓN":"Acondicionar boca de pozo + herramientas de v/b.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"254","TIEMPO":""
            })

            # 22) BAJA VARILLAS
            m.append({
                "MANIOBRA NORMALIZADA":"BAJA VARILLAS","PUNTO PROGRAMA":22,
                "DESCRIPCIÓN":"Tareas varias durante bajado de varillas.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"255","TIEMPO":""
            })

            # —14) BAJA VARILLAS EN SIMPLE (punto 14)
            parts = []
            # iteramos varillas_fin en orden inverso
            for _, row in varillas_fin.iloc[::-1].iterrows():
                # cantidad
                qty = int(row["CANTIDAD"])  
                # elemento y condición
                elem = row["ELEMENTO"]
                cond = row["CONDICIÓN"] if pd.notnull(row["CONDICIÓN"]) else ""
                # profundidad en mts
                depth = f" {int(row['PROFUNDIDAD'])} mts" if pd.notnull(row["PROFUNDIDAD"]) else ""
                # diámetro
                diam = f" de {row['DIÁMETRO']}" if pd.notnull(row["DIÁMETRO"]) else ""
                # acero V/B
                acero_vb = f", {row['ACERO V/B']}" if pd.notnull(row["ACERO V/B"]) else ""
                # cupla SH/FS
                cupla = f", cupla {row['CUPLA SH/FS']}" if pd.notnull(row["CUPLA SH/FS"]) else ""
                # acero cupla
                acero_cupla = f", {row['ACERO CUPLA']}" if pd.notnull(row["ACERO CUPLA"]) else ""
                # comentario (solo para marcar BAJA EN SIMPLE)
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""
                
                # armamos la pieza
                piece = f"{qty} {elem}"
                if cond:
                    piece += f" {cond}"
                piece += depth + diam + acero_vb + cupla + acero_cupla + comm_tag

                parts.append(piece.strip())

            diseño = " + ".join(parts)

            m.append({
                "MANIOBRA NORMALIZADA":   "BAJA VARILLAS EN SIMPLE",
                "PUNTO PROGRAMA":         23,
                "DESCRIPCIÓN": (
                    "Tomar datos de bomba y bajarla + sarta de v/b limpiando todas las conexiones "
                    "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                    "de acero de varillas. Diseño a bajar: " + diseño + "."
                ),
                "ACTIVITY_PHASE": "SP04",
                "ACTIVITY_CODE":  "SP16",
                "ACTIVITY_SUBCODE":"255",
                "TIEMPO": ""
            })

            # 24) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH","PUNTO PROGRAMA":24,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })

            # 25) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD","PUNTO PROGRAMA":25,
                "DESCRIPCIÓN":"Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })

            # 26) VARIOS
            m.append({
                "MANIOBRA NORMALIZADA":"VARIOS","PUNTO PROGRAMA":26,
                "DESCRIPCIÓN":"Tareas varias.","ACTIVITY_PHASE":"SPV","ACTIVITY_CODE":"SPV","ACTIVITY_SUBCODE":"SPV","TIEMPO":""
            })

            # 27) ARMA BDP
            m.append({
                "MANIOBRA NORMALIZADA":"ARMA BDP","PUNTO PROGRAMA":27,
                "DESCRIPCIÓN":"Armar BDP.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP15","ACTIVITY_SUBCODE":"260","TIEMPO":""
            })

            # 28) DESMONTA EQUIPO
            m.append({
                "MANIOBRA NORMALIZADA":"DESMONTA EQUIPO","PUNTO PROGRAMA":28,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado.","ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP11","ACTIVITY_SUBCODE":"202","TIEMPO":""
            })

            # finalmente, muestro y guardo
            df_prog = pd.DataFrame(m)
            display(df_prog)
            df_prog.to_excel("cambio_bba_tubing_con_ancla_program.xlsx", index=False)
            print("✅ Guardado: cambio_bba_tubing_con_ancla_program.xlsx")

        # ——————————————————————————————
        def generar_cambio_bba_tubing_con_ancla():
            global df0, meta_pos, varillas_act, tubing_act, tubing_fin

            # 0) recojo inputs
            meta = {k: df0.iat[r,c] for k,(r,c) in meta_pos.items()}

            # 1) construyo maniobras 1–17
            m = _build_base_maniobras(meta, varillas_act, tubing_act)

            # 2) intento obtener prof_m desde sheet / zapato / bomba
            prof_m = None
            ancla_row = tubing_fin[
                tubing_fin["ELEMENTO"].str.contains("ANCLA", case=False, na=False)
            ]

            # ya no levantamos error aquí, dejamos que prof_m quede None si no existe
            if not ancla_row.empty:
                raw = ancla_row["PROFUNDIDAD"].iloc[0]
                if pd.notna(raw) and raw > 0:
                    prof_m = raw
                else:
                    # probar ZAPATO
                    zap = tubing_fin[
                        tubing_fin["ELEMENTO"].str.contains("ZAPATO", case=False, na=False)
                    ]
                    if not zap.empty and pd.notna(zap["PROFUNDIDAD"].iloc[0]):
                        prof_m = zap["PROFUNDIDAD"].iloc[0]
                    else:
                        # probar BOMBA
                        bom = tubing_fin[
                            tubing_fin["ELEMENTO"].str.contains("BOMBA", case=False, na=False)
                        ]
                        if not bom.empty and pd.notna(bom["PROFUNDIDAD"].iloc[0]):
                            prof_m = bom["PROFUNDIDAD"].iloc[0]

            # 3) si no hay prof_m válida, pido con widgets y espero OK
            if prof_m is None or prof_m <= 0:
                display(widgets.HTML(
                    "<span style='color: darkorange; font-weight: bold;'>"
                    "No se encontró ANCLA con profundidad válida.<br>"
                    "Por favor ingresa manualmente la profundidad de ancla (m):"
                    "</span>"
                ))
                depth_input = widgets.FloatText(
                    description="Profundidad ANCLA (m):",
                    placeholder="ej. 1800"
                )
                ok_button = widgets.Button(description="OK", button_style="primary")
                display(widgets.HBox([depth_input, ok_button]))

                def _on_ok(b):
                    clear_output()
                    val = depth_input.value
                    if val and val > 0:
                        _añadir_fija_y_finalizar(m, tubing_fin, val)
                    else:
                        display(HTML(
                            "<span style='color: red;'>"
                            "Debe ingresar un número mayor que 0.</span>"
                        ))
                ok_button.on_click(_on_ok)

            else:
                # si ya tenemos prof_m, continuamos normalmente
                _añadir_fija_y_finalizar(m, tubing_fin, prof_m)

        def generar_cambio_bba_tubing_sin_ancla():
            # 0) inputs
            meta = {k: df0.iat[r,c] for k,(r,c) in meta_pos.items()}

            # 1) maneja 1–12 igual que en _build_base_maniobras()
            m = _build_base_maniobras(meta, varillas_act, tubing_act)

            # 2) Reemplaza la maniobra “SACA TUBING EN DOBLE” (punto 13):
            for man in m:
                if man["MANIOBRA NORMALIZADA"] == "SACA TUBING EN DOBLE":
                    # Construir "diseño_act" evaluando conn por cada fila de tubing_act
                    diseño_act = " + ".join(
                        f"{int(row['CANTIDAD'])} {row['ELEMENTO']} {row['DIÁMETRO']}in("
                        + ("DOBLE" if "DOBLE" in str(row["COMENTARIO"]).upper() else "SIMPLE")
                        + ")"
                        for _, row in tubing_act.iterrows()
                    )

                    man.update({
                        "PUNTO PROGRAMA": 13,
                        "DESCRIPCIÓN": (
                            
                            "Sacar columna, buscando pérdida y completando pozo. "
                            "Desarmar componentes que requieran reemplazo. "
                            f"Diseño a extraer: {diseño_act}. "
                            "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, "
                            "desgaste, incrustaciones o sobretorque. Registrar evidencia fotográfica del estado del material. "
                            "Asentar en OW grado de acero del material extraído. "
                            "Indicar si evidencia falla visible. Asentar número de bomba y estado de cabezal y filtro."
                        )
                    })
                    break   # solo una

            # 3) Quita cualquier “FIJA ANCLA” (si viniera de la versión con ancla)
            m = [man for man in m if man["MANIOBRA NORMALIZADA"]!="FIJA ANCLA"]

            # 4) Añade del 19 en adelante tal cual en _añadir_fija_y_finalizar(), **pero sin el 18**:
            
            # — 21) ACONDICIONAMIENTO PARA BAJAR VARILLAS
            m.append({ "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA BAJAR VARILLAS","PUNTO PROGRAMA":21,
                       "DESCRIPCIÓN":"Acondicionar boca de pozo + herramientas de v/b.",
                       "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"254","TIEMPO":"" })
            # — 22) BAJA VARILLAS
            m.append({ "MANIOBRA NORMALIZADA":"BAJA VARILLAS","PUNTO PROGRAMA":22,
                       "DESCRIPCIÓN":"Tareas varias durante bajado de varillas.",
                       "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"255","TIEMPO":"" })
            # —14) BAJA VARILLAS EN SIMPLE (punto 14)
            parts = []
            # iteramos varillas_fin en orden inverso
            for _, row in varillas_fin.iloc[::-1].iterrows():
                # cantidad (si no está, asumimos 1)
                qty = int(row["CANTIDAD"]) if pd.notnull(row["CANTIDAD"]) else 1
                # elemento
                elem = row["ELEMENTO"]
                # condición (si existe)
                cond = f" {row['CONDICIÓN']}" if pd.notnull(row["CONDICIÓN"]) else ""
                # diámetro (si existe)
                diam = f" {row['DIÁMETRO']}" if pd.notnull(row["DIÁMETRO"]) else ""
                # acero V/B
                acero_vb = f", {row['ACERO V/B']}" if pd.notnull(row["ACERO V/B"]) else ""
                # cupla SH/FS
                cupla = f", cupla {row['CUPLA SH/FS']}" if pd.notnull(row["CUPLA SH/FS"]) else ""
                # acero cupla
                acero_cupla = f", {row['ACERO CUPLA']}" if pd.notnull(row["ACERO CUPLA"]) else ""
                # comentario para etiqueta de baja
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                # armamos la pieza
                piece = f"{qty} {elem}{cond}{diam}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño = " + ".join(parts)

            m.append({
                "MANIOBRA NORMALIZADA":   "BAJA VARILLAS EN SIMPLE",
                "PUNTO PROGRAMA":         23,
                "DESCRIPCIÓN": (
                    "Tomar datos de bomba y bajarla + sarta de v/b limpiando todas las conexiones "
                    "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                    "de acero de varillas. Diseño a bajar: " + diseño + "."
                ),
                "ACTIVITY_PHASE": "SP04",
                "ACTIVITY_CODE":  "SP16",
                "ACTIVITY_SUBCODE":"255",
                "TIEMPO": ""
            })


            # 24) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH","PUNTO PROGRAMA":24,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })

            # 25) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD","PUNTO PROGRAMA":25,
                "DESCRIPCIÓN":"Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })

            # 26) VARIOS
            m.append({
                "MANIOBRA NORMALIZADA":"VARIOS","PUNTO PROGRAMA":26,
                "DESCRIPCIÓN":"Tareas varias.","ACTIVITY_PHASE":"SPV","ACTIVITY_CODE":"SPV","ACTIVITY_SUBCODE":"SPV","TIEMPO":""
            })

            # 27) ARMA BDP
            m.append({
                "MANIOBRA NORMALIZADA":"ARMA BDP","PUNTO PROGRAMA":27,
                "DESCRIPCIÓN":"Armar BDP.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP15","ACTIVITY_SUBCODE":"260","TIEMPO":""
            })

            # 28) DESMONTA EQUIPO
            m.append({
                "MANIOBRA NORMALIZADA":"DESMONTA EQUIPO","PUNTO PROGRAMA":28,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado.","ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP11","ACTIVITY_SUBCODE":"202","TIEMPO":""
            })

            df_sin_ancla = pd.DataFrame(m)
            display(df_sin_ancla)
            df_sin_ancla.to_excel("programa_sin_ancla_b_program.xlsx", index=False)
            print("✅ Guardado: programa_sin_ancla_b.xlsx")

        def generar_mov_varillas_ph_intermedia():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            # — Constantes de cálculo de ancla —
            WALL_THICKNESS       = {2.875: 0.434, 3.5: 0.508}
            CONVERSION_PIES_METROS = 3.28084
            COEF_POISSON         = 0.3
            MODULO_YOUNG         = 30_000_000
            COEF_EXPANSION       = 0.0000069
            GRADIENTE_FLUIDO     = 0.5
            TEMP_SUPERFICIE_C    = 30
            TEMP_MEDIA_C         = 12
            NIVEL_ESTATICO_PIE   = 656

            # 0) Inputs
            meta        = {k: df0.iat[r, c] for k,(r,c) in meta_pos.items()}


            # 1) Base maniobras 1–10
            m = _build_base_maniobras(meta, varillas_act, tubing_act)

            # 2) Inserto PH intermedia tras “SACA VARILLAS EN DOBLE”
            idx = next(i for i,row in enumerate(m)
                       if row["MANIOBRA NORMALIZADA"]=="SACA VARILLAS EN SIMPLE") + 1

            m.insert(idx, {
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "PUNTO PROGRAMA":     None,
                "DESCRIPCIÓN":(
                    "Acondicionar superficie para realizar prueba hidráulica intermedia "
                    
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })
            m.insert(idx+1, {
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "PUNTO PROGRAMA":     None,
                "DESCRIPCIÓN":(
                    "Realizar PH intermedia sobre probador con 1000 psi. "
                    "Si es negativa, detener,revisar e informar a sala mtr; si es positiva, continuar programa."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })

            # 3) Renumero puntos 1–12
            for i, maniobra in enumerate(m, start=1):
                maniobra["PUNTO PROGRAMA"] = i

            # 4) A partir de aquí, maniobras 13–30
            siguiente = len(m) + 1

            
                    
            
            # — Validación de PROFUNDIDAD de ANCLA —
            
            ancla = tubing_fin[tubing_fin["ELEMENTO"].str.contains("ANCLA", case=False, na=False)]
            if ancla.empty or pd.isna(ancla["PROFUNDIDAD"].iloc[0]):
                # Aviso de que faltan datos, pero no abortamos el proceso
                print("❌ Falta la profundidad de ancla. No se calculó tensión ni estiramiento.")
                desc10 = "Falta la profundidad de ancla; no se calculó tensión ni estiramiento."
            else:
                prof_m = float(ancla["PROFUNDIDAD"].iloc[0])

                # 20) FIJA ANCLA
                #diam_ext = float(ancla["DIÁMETRO"].iloc[0])
                diam_ext =2.875
                thickness = WALL_THICKNESS.get(diam_ext, 0.434)

                prof_ft   = prof_m * CONVERSION_PIES_METROS
                dynamic_m = max(prof_m - (NIVEL_ESTATICO_PIE/CONVERSION_PIES_METROS), 0)
                dynamic_ft= dynamic_m * CONVERSION_PIES_METROS

                thickness = WALL_THICKNESS.get(diam_ext, 0.434)
                diam_int  = diam_ext - thickness

                # ratio para F1
                ratio = ((COEF_POISSON * dynamic_ft / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                F1    = 6.4918 * dynamic_ft * GRADIENTE_FLUIDO * ratio
                seccion_pared = 1.812
                F2    = MODULO_YOUNG * COEF_EXPANSION * ((TEMP_SUPERFICIE_C - TEMP_MEDIA_C)/2) * seccion_pared

                B     = ((COEF_POISSON * NIVEL_ESTATICO_PIE / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                A     = 6.4918 * GRADIENTE_FLUIDO * ((seccion_pared / (diam_ext**2)) * NIVEL_ESTATICO_PIE)
                F3    = A * B

                tension_total   = F1 + F2 - F3
                estiramiento_in = 0.22 * (prof_ft/1000) * (tension_total/1000)
                est_cm          = estiramiento_in * 2.54
            
                
                desc10 = f"Fijar ancla con {tension_total:.2f} lbs y {est_cm:.2f} cm de estiramiento."
            m.append({
                "MANIOBRA NORMALIZADA": "FIJA ANCLA",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP17",
                "ACTIVITY_SUBCODE":     "211",
                "PUNTO PROGRAMA":       10,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc10
            })
            
            # 23) ACONDICIONAMIENTO PARA BAJAR VARILLAS
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "PUNTO PROGRAMA":       siguiente,
                "DESCRIPCIÓN":"Acondicionar boca de pozo + herramientas de v/b.",
                "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"254","TIEMPO":""
            })
            siguiente += 1

            # 24) BAJA VARILLAS
            m.append({
                "MANIOBRA NORMALIZADA":"BAJA VARILLAS",
                "PUNTO PROGRAMA":       siguiente,
                "DESCRIPCIÓN":"Tareas varias durante bajado de varillas.",
                "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"255","TIEMPO":""
            })
            siguiente += 1

            # 25) BAJA VARILLAS EN DOBLE (diseño de instalación final)
            parts_fin_vb = [
                f"{int(r['CANTIDAD'])} {r['ELEMENTO']} {r['DIÁMETRO']}in"
                for _, r in varillas_fin.iterrows()
            ]
            diseño_fin_vb = " + ".join(parts_fin_vb)
            m.append({
                "MANIOBRA NORMALIZADA":"BAJA VARILLAS EN DOBLE",
                "PUNTO PROGRAMA":       siguiente,
                "DESCRIPCIÓN":(
                    "Tomar datos de bomba y bajarla + sarta de v/b limpiando todas las conexiones con detergente biodegradable. "
                    f"Diseño a bajar: {diseño_fin_vb}."
                ),
                "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP16","ACTIVITY_SUBCODE":"255","TIEMPO":""
            })
            siguiente += 1

            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "PUNTO PROGRAMA":       siguiente,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })
            siguiente += 1

            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "PUNTO PROGRAMA":       siguiente,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })
            siguiente += 1

            # 28) VARIOS
            m.append({
                "MANIOBRA NORMALIZADA":"VARIOS",
                "PUNTO PROGRAMA":       siguiente,
                "DESCRIPCIÓN":"Tareas varias.",
                "ACTIVITY_PHASE":"SPV","ACTIVITY_CODE":"SPV","ACTIVITY_SUBCODE":"SPV","TIEMPO":""
            })
            siguiente += 1

            # 29) ARMA BDP
            m.append({
                "MANIOBRA NORMALIZADA":"ARMA BDP",
                "PUNTO PROGRAMA":       siguiente,
                "DESCRIPCIÓN":"Armar BDP.",
                "ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP15","ACTIVITY_SUBCODE":"260","TIEMPO":""
            })
            siguiente += 1

            # 30) DESMONTA EQUIPO
            m.append({
                "MANIOBRA NORMALIZADA":"DESMONTA EQUIPO",
                "PUNTO PROGRAMA":       siguiente,
                "DESCRIPCIÓN":(
                    "Acondicionar boca de pozo, material sobrante y locación "
                    "y accesorios de superficie. Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling "
                    "la finalización de la intervención y transporte a próxima locación. Generar acta de entrega/recepción "
                    "de locación. Indicar si el puente de producción queda armado."
                ),
                "ACTIVITY_PHASE":"SP01","ACTIVITY_CODE":"SP11","ACTIVITY_SUBCODE":"202","TIEMPO":""
            })



            # 5) Muestro y guardo
            df_prog = pd.DataFrame(m)
            display(df_prog)
            df_prog.to_excel("mov_varillas_ph_intermedia_program.xlsx", index=False)
            print("✅ Guardado: mov_varillas_ph_intermedia_program.xlsx")

        def generar_cambio_Pesca_y_enganche_x_bm():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            # — Constantes de cálculo de ancla —
            WALL_THICKNESS        = {2.875: 0.434, 3.5: 0.508}
            CONVERSION_PIES_METROS = 3.28084
            COEF_POISSON          = 0.3
            MODULO_YOUNG          = 30_000_000
            COEF_EXPANSION        = 0.0000069
            GRADIENTE_FLUIDO      = 0.5
            TEMP_SUPERFICIE_C     = 30
            TEMP_MEDIA_C          = 12
            NIVEL_ESTATICO_PIE    = 656

            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo}. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

            # Maniobra 3: DESARMA BDP
            desc3 = "Desarmar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BDP",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP34",
                "ACTIVITY_SUBCODE":     "210",
                "Punto programa":       3,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc3
            })

            # — 6) ACONDICIONA PARA SACAR VARILLAS (con PH inicial)
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","Punto programa":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. Tomar peso de sarta, y registrar en OW. Realizar PH inicial con 1000 psi e informar resultados, de ser positiva continuar con maniobras.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
                        
            # —11) SACA VARILLAS EN PESCA
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS EN PESCA EN DOBLE",
                "Punto programa":7,
                "DESCRIPCIÓN":(
                    "Saca varillas en doble hasta punto de pesca, completando pozo. Desarmar componentes que requieran reemplazo. "
                    "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                    "Registrar evidencia fotográfica del estado del material y punto de pesca. Asentar en OW grado de acero del material extraído."
                ),               
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""
            })
            
            
             # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm  = row["COMENTARIO"] or ""
                # convierto a float para comparar
                try:
                    dia_val = float(dia)
                except:
                    dia_val = None

                # 1) lógica especial para PESO
                if "PESO" in elem.upper() and dia_val is not None:
                    if abs(dia_val - 1.5) < 1e-3:
                        modo = "(simple)"
                    elif abs(dia_val - 1.125) < 1e-3:
                        modo = "(doble)"
                    else:
                        # cae al modo por defecto
                        modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                else:
                    # 2) modo por comentario
                    modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""

                parts.append(f"{qty} {elem} {dia}in {modo}".strip())

            diseño = " + ".join(parts)
                

            desc_sv_simple = (
                "Bajar sarta de varillas en doble + pescador y Pescar. Desclavar bomba. "f"Diseño a extraer: {diseño}. "
                "Sacar varillas pescadas, desarmando conexión, completando pozo. "
                "Desarmar componentes que requieran reemplazo. Reemplazar 5 varillas por arriba y por debajo en zona de pesca."
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "Punto programa":        10,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })
            
            # — 8) CIRCULA
            m.append({
                "MANIOBRA NORMALIZADA":"CIRCULA","Punto programa":8,
                "DESCRIPCIÓN":"Circular pozo, 1.5 veces la capacidad de tubing por directa hasta retorno limpio para asegurar limpieza de los materiales extraídos. Si no se observa circulación informar si es por punta de instalación obstruida o porque el pozo admite.",
                "ACTIVITY_PHASE":"SP05","ACTIVITY_CODE":"SP20","ACTIVITY_SUBCODE":"218","TIEMPO":""
            })
            
                        
            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo + herramientas de v/b."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })

            # Maniobra 14: BAJA VARILLAS
            desc14 = "Tareas varias durante bajada de varillas."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       14,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc14
            })

           # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de bomba y bajarla + sarta de v/b  limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       15,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc15
            })


            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "Punto programa":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "Punto programa":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })



            # Maniobra 16: VARIOS
            desc16 = "Tareas varias."
            m.append({
                "MANIOBRA NORMALIZADA": "VARIOS",
                "ACTIVITY_PHASE":       "SPV",
                "ACTIVITY_CODE":        "SPV",
                "ACTIVITY_SUBCODE":     "SPV",
                "Punto programa":       16,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc16
            })

            # Maniobra 17: ARMA BDP
            desc17 = "Armar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "ARMA BDP",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP15",
                "ACTIVITY_SUBCODE":     "260",
                "Punto programa":       17,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc17
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("cambio_Pesca y enganche_x_bm_program.xlsx", index=False)
            print("\n✅ Archivo guardado como 'cambio_Pesca y enganche_x_bm_program.xlsx'")
        
        def generar_cambio_bm():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            
            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo}. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

            # Maniobra 3: DESARMA BDP
            desc3 = "Desarmar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BDP",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP34",
                "ACTIVITY_SUBCODE":     "210",
                "Punto programa":       3,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc3
            })

            # — 6) ACONDICIONA PARA SACAR VARILLAS (con PH inicial)
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","Punto programa":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. Tomar peso de sarta, y registrar en OW. Realizar PH inicial con 1000 psi e informar resultados, de ser positiva continuar con maniobras.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
                        
            # — 7) SACA VARILLAS (solo peso/anomalías)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS","Punto programa":7,
                "DESCRIPCIÓN":"Tareas varias durante saca de varillas",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""
            })
               
                        
             # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm  = row["COMENTARIO"] or ""
                # convierto a float para comparar
                try:
                    dia_val = float(dia)
                except:
                    dia_val = None

                # 1) lógica especial para PESO
                if "PESO" in elem.upper() and dia_val is not None:
                    if abs(dia_val - 1.5) < 1e-3:
                        modo = "(simple)"
                    elif abs(dia_val - 1.125) < 1e-3:
                        modo = "(doble)"
                    else:
                        # cae al modo por defecto
                        modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                else:
                    # 2) modo por comentario
                    modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""

                parts.append(f"{qty} {elem} {dia}in {modo}".strip())

            diseño = " + ".join(parts)

            desc_sv_simple = (
                "Desclavar bomba"
                "Sacar sarta desarmando, completando pozo. Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído."
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "Punto programa":        10,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })
            
            # — 8) CIRCULA
            m.append({
                "MANIOBRA NORMALIZADA":"CIRCULA","Punto programa":8,
                "DESCRIPCIÓN":"Circular pozo por directa hasta retorno.Si no se observa circulación informar si es por punta de instalación obstruida o porque el pozo admite.",
                "ACTIVITY_PHASE":"SP05","ACTIVITY_CODE":"SP20","ACTIVITY_SUBCODE":"218","TIEMPO":""
            })
            
            
            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo + herramientas de v/b."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })

            # Maniobra 14: BAJA VARILLAS
            desc14 = "Tareas varias durante bajada de varillas."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       14,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc14
            })

           # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de bomba y bajarla + sarta de v/b  limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       15,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc15
            })


            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "Punto programa":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "Punto programa":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })



            # Maniobra 16: VARIOS
            desc16 = "Tareas varias."
            m.append({
                "MANIOBRA NORMALIZADA": "VARIOS",
                "ACTIVITY_PHASE":       "SPV",
                "ACTIVITY_CODE":        "SPV",
                "ACTIVITY_SUBCODE":     "SPV",
                "Punto programa":       16,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc16
            })

            # Maniobra 17: ARMA BDP
            desc17 = "Realizar final sobre bomba con 1000 psi. Realizar prueba de bomba. Si la prueba es deficiente informar a Supervisor de Pulling."
            m.append({
                "MANIOBRA NORMALIZADA": "ARMA BDP",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP15",
                "ACTIVITY_SUBCODE":     "260",
                "Punto programa":       17,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc17
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("generar_cambio_bm_program.xlsx", index=False)
            print("\n✅ Archivo guardado como 'generar_cambio_bm_program.xlsx'")

        def generar_bayler_bm():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            
            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo}. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

            # Maniobra 3: DESARMA BDP
            desc3 = "Desarmar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BDP",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP34",
                "ACTIVITY_SUBCODE":     "210",
                "Punto programa":       3,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc3
            })

            # — 6) ACONDICIONA PARA SACAR VARILLAS (con PH inicial)
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","Punto programa":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. Tomar peso de sarta, y registrar en OW. Realizar PH inicial con 1000 psi e informar resultados, de ser positiva continuar con maniobras.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
                        
            # — 7) SACA VARILLAS (solo peso/anomalías)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS","Punto programa":7,
                "DESCRIPCIÓN":"Registrar peso de la sarta y eventuales anomalías antes de iniciar la extracción.Tareas varias durante saca de varillas",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""})
            
             # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm  = row["COMENTARIO"] or ""
                # convierto a float para comparar
                try:
                    dia_val = float(dia)
                except:
                    dia_val = None

                # 1) lógica especial para PESO
                if "PESO" in elem.upper() and dia_val is not None:
                    if abs(dia_val - 1.5) < 1e-3:
                        modo = "(simple)"
                    elif abs(dia_val - 1.125) < 1e-3:
                        modo = "(doble)"
                    else:
                        # cae al modo por defecto
                        modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                else:
                    # 2) modo por comentario
                    modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""

                parts.append(f"{qty} {elem} {dia}in {modo}".strip())

            diseño = " + ".join(parts)

            desc_sv_simple = (
                "Desclavar bomba."
                "Sacar sarta desarmando, completando pozo. Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído."
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "Punto programa":        10,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })
            
            # 11) ACONDICIONAMIENTO PARA SACAR CAÑOS
            m.append({ "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA SACAR CAÑOS","Punto programa":11,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, completar con ASDF, desempaquetar y montar conjunto BOP.Montar piso de trabajo.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"252","TIEMPO":"" })

            # 12) SACA TUBING
            m.append({ "MANIOBRA NORMALIZADA":"SACA TUBING","Punto programa":12,
                "DESCRIPCIÓN":"Tareas generales durante la sacada de tubing.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })

            # 13) SACA TUBING EN DOBLE with anchor
            parts_tub = [
                f"{int(r['CANTIDAD'])} {r['ELEMENTO']} {r['DIÁMETRO']}in "
                f"({'DOBLE' if 'DOBLE' in str(r['COMENTARIO']).upper() else 'SIMPLE'})"
                for _,r in tubing_act.iterrows()
            ]
            diseño_tub = " + ".join(parts_tub)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA TUBING EN DOBLE","Punto programa":13,
                "DESCRIPCIÓN":(
                    "Librar ANCLA.Sacar columna en tiro doble, completando pozo. Desarmar componentes que requieran reemplazo. "
                    f"Diseño a extraer: {diseño_tub}. Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído. Solicitar envío de bomba a taller de Pecom de inmediato para desarmar e inspeccionar. Indicar si evidencia falla visible. Asentar número de bomba y estado de cabezal y filtro.."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })    
            
            # 12) ARMA BAYLER
            m.append({ "MANIOBRA NORMALIZADA":"ARMA BAYLER","Punto programa":12,
                "DESCRIPCIÓN":"Armar Bayler.","ACTIVITY_PHASE":"SP04","ACTIVITY_CODE":"SP15","ACTIVITY_SUBCODE":"209","TIEMPO":"" })
            
            # 17) BAJA TUBING EN SIMPLE: Construyo “INSTALACIÓN FINAL tubing” a partir de tubing_fin
            partes_fin_tub = []
            for _, r in tubing_act.iterrows():
                # Sólo si hay un ELEMENTO válido
                if isinstance(r["ELEMENTO"], str) and r["ELEMENTO"].strip():
                    cant = int(r["CANTIDAD"]) if pd.notna(r["CANTIDAD"]) else 0
                    elem = r["ELEMENTO"].strip()
                    # Armo únicamente con cantidad y elemento
                    partes_fin_tub.append(f"{cant} {elem}")

            # Invierto el orden
            partes_fin_tub.reverse()

            # Concateno con " + " y siempre empiezo con "CONJUNTO BAYLER"
            if partes_fin_tub:
                diseño_fin_tub = "CONJUNTO BAYLER + " + " + ".join(partes_fin_tub)
            else:
                diseño_fin_tub = "CONJUNTO BAYLER"

            desc9 = (
                "Bajar conjunto bayler + tubing en doble. "
                f"Diseño a bajar: {diseño_fin_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar número de bomba bajado. Limpiar e informar metros de avance y retorno"
            )
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING EN SIMPLE",
                "Punto programa":       17,
                "DESCRIPCIÓN":          desc9,
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "TIEMPO":               "",
            })    
            
            # 13) SACA TUBING EN DOBLE  (sin incluir diámetro ni comentario)
            parts_tub = []
            for _, r in tubing_act.iterrows():
                if isinstance(r["ELEMENTO"], str) and r["ELEMENTO"].strip():
                    cant = int(r["CANTIDAD"]) if pd.notna(r["CANTIDAD"]) else 0
                    elem = r["ELEMENTO"].strip()
                    # Solo concateno cantidad y elemento:
                    parts_tub.append(f"{cant} {elem}")

            # Concateno con " + "
            diseño_tub = " + ".join(parts_tub)

            m.append({
                "MANIOBRA NORMALIZADA": "SACA TUBING EN DOBLE",
                "Punto programa":       13,
                "DESCRIPCIÓN": (
                    "Sacar tubing, completando pozo. Desarmar componentes que requieran reemplazo. "
                    f"Diseño a extraer: {diseño_tub}."
                ),
                "ACTIVITY_PHASE":   "SP03",
                "ACTIVITY_CODE":    "SP24",
                "ACTIVITY_SUBCODE": "253",
                "TIEMPO":           ""
            })    
            
            # 12) DESARMA BAYLER
            m.append({ "MANIOBRA NORMALIZADA":"ARMA BAYLER","Punto programa":12,
                "DESCRIPCIÓN":"Desarmar Bayler.Asentar en OW observaciones significativas en cuanto a eventual presencia de incrustaciones, sólidos. Registrar evidencia fotográfica y tomar muestra si así lo requiere el ingeniero.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP25","ACTIVITY_SUBCODE":"229","TIEMPO":"" })
            
            
            # Maniobra 9: BAJA TUBING EN SIMPLE
            # Construyo “INSTALACIÓN FINAL tubing” a partir de tubing_fin
            
            
            partes_fin_tub = []
            for _, r in tubing_fin.iterrows():
                elem = r["ELEMENTO"]
                if not (isinstance(elem, str) and elem.strip()):
                    continue

                tokens = []
                # cantidad
                if pd.notna(r["CANTIDAD"]):
                    tokens.append(str(int(r["CANTIDAD"])))
                # elemento
                tokens.append(elem.strip())
                # condición
                if pd.notna(r["CONDICIÓN"]):
                    tokens.append(r["CONDICIÓN"].strip())
                # diámetro
                if pd.notna(r["DIÁMETRO"]):
                    tokens.append(f"{r['DIÁMETRO']}in")
                # profundidad
                if pd.notna(r["PROFUNDIDAD"]):
                    tokens.append(f"en {r['PROFUNDIDAD']}m")

                parte = " ".join(tokens)
                # comentario libre
                if pd.notna(r["COMENTARIO"]) and r["COMENTARIO"].strip():
                    parte += f" ({r['COMENTARIO'].strip()})"

                partes_fin_tub.append(parte)
            

            # Invierto el orden
            partes_fin_tub.reverse()

            # Concateno con " + "
            diseño_fin_tub = " + ".join(partes_fin_tub) if partes_fin_tub else ""

            desc9 = (
                "Profundizar columna de TBG calibrando, midiendo, limpiando, engrasando y torqueando las conexiones. "
                "Bajar instalación en simple. "
                f"Diseño a bajar: {diseño_fin_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar número de bomba bajado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "Punto programa":       9,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc9
            })

            # Maniobra 11: ACONDICIONA PARA PH
            desc11 = "Acondicionar superficie para realizar prueba hidráulica."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONA PARA PH",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "259",
                "Punto programa":       11,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc11
            })

            # Maniobra 12: PRUEBA DE HERMETICIDAD
            desc12 = "Realizar PH inicial, intermedia y final con 800, 900 y 1000 psi respectivamente.Registrar la misma en OpenWells.. Librar SO con 2000 psi.Si la prueba es deficiente informar a Supervisor de Pulling."
            m.append({
                "MANIOBRA NORMALIZADA": "PRUEBA DE HERMETICIDAD",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "205",
                "Punto programa":       12,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc12
            })    
                
            # 20) FIJA ANCLA
            # — Constantes de cálculo de ancla —
            WALL_THICKNESS        = {2.875: 0.434, 3.5: 0.508}
            CONVERSION_PIES_METROS = 3.28084
            COEF_POISSON          = 0.3
            MODULO_YOUNG          = 30_000_000
            COEF_EXPANSION        = 0.0000069
            GRADIENTE_FLUIDO      = 0.5
            TEMP_SUPERFICIE_C     = 30
            TEMP_MEDIA_C          = 12
            NIVEL_ESTATICO_PIE    = 656
            
            # — Validación de PROFUNDIDAD de ANCLA —
            ancla = tubing_fin[tubing_fin["ELEMENTO"].str.contains("ANCLA", case=False, na=False)]
            if ancla.empty or pd.isna(ancla["PROFUNDIDAD"].iloc[0]):
                # Aviso de que faltan datos, pero no abortamos el proceso
                print("❌ Falta la profundidad de ancla. No se calculó tensión ni estiramiento.")
                desc10 = "Falta la profundidad de ancla; no se calculó tensión ni estiramiento."
            else:
                prof_m = float(ancla["PROFUNDIDAD"].iloc[0])

                # 20) FIJA ANCLA
                diam_ext = float(ancla["DIÁMETRO"].iloc[0])
                thickness = WALL_THICKNESS.get(diam_ext, 0.434)

                prof_ft   = prof_m * CONVERSION_PIES_METROS
                dynamic_m = max(prof_m - (NIVEL_ESTATICO_PIE/CONVERSION_PIES_METROS), 0)
                dynamic_ft= dynamic_m * CONVERSION_PIES_METROS

                thickness = WALL_THICKNESS.get(diam_ext, 0.434)
                diam_int  = diam_ext - thickness

                # ratio para F1
                ratio = ((COEF_POISSON * dynamic_ft / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                F1    = 6.4918 * dynamic_ft * GRADIENTE_FLUIDO * ratio
                seccion_pared = 1.812
                F2    = MODULO_YOUNG * COEF_EXPANSION * ((TEMP_SUPERFICIE_C - TEMP_MEDIA_C)/2) * seccion_pared

                B     = ((COEF_POISSON * NIVEL_ESTATICO_PIE / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                A     = 6.4918 * GRADIENTE_FLUIDO * ((seccion_pared / (diam_ext**2)) * NIVEL_ESTATICO_PIE)
                F3    = A * B

                tension_total   = F1 + F2 - F3
                estiramiento_in = 0.22 * (prof_ft/1000) * (tension_total/1000)
                est_cm          = estiramiento_in * 2.54

                desc10 = f"Fijar ancla con {tension_total:.2f} lbs y {est_cm:.2f} cm de estiramiento."
            m.append({
                "MANIOBRA NORMALIZADA": "FIJA ANCLA",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP17",
                "ACTIVITY_SUBCODE":     "211",
                "Punto programa":       10,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc10
            })    
                
            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo, montar piso de trabajo + herramientas de v/b.."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })

            # Maniobra 14: BAJA VARILLAS
            desc14 = "Tareas varias durante bajada de varillas."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       14,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc14
            })

           # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de bomba y bajarla + sarta de v/b  limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       15,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc15
            })


            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "Punto programa":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "Punto programa":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi.Realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })



            # Maniobra 16: VARIOS
            desc16 = "Tareas varias."
            m.append({
                "MANIOBRA NORMALIZADA": "VARIOS",
                "ACTIVITY_PHASE":       "SPV",
                "ACTIVITY_CODE":        "SPV",
                "ACTIVITY_SUBCODE":     "SPV",
                "Punto programa":       16,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc16
            })

            # Maniobra 17: ARMA BDP
            desc17 = "Arma BDP"
            m.append({
                "MANIOBRA NORMALIZADA": "ARMA BDP",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP15",
                "ACTIVITY_SUBCODE":     "260",
                "Punto programa":       17,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc17
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("generar_bayler_bm_program.xlsx", index=False)
            print("\n✅ Archivo guardado como 'generar_bayler_bm_program.xlsx'")
        
        def generar_Pesca_Cambio_bba_tub_bm():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            
            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo}. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

            # Maniobra 3: DESARMA BDP
            desc3 = "Desarmar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BDP",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP34",
                "ACTIVITY_SUBCODE":     "210",
                "Punto programa":       3,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc3
            })

            # — 6) ACONDICIONA PARA SACAR VARILLAS (con PH inicial)
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","Punto programa":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. Tomar peso de sarta, y registrar en OW. Realizar PH inicial con 1000 psi e informar resultados, de ser positiva continuar con maniobras.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
                        
            # —11) SACA VARILLAS EN PESCA
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS EN PESCA EN DOBLE",
                "Punto programa":7,
                "DESCRIPCIÓN":(
                    "Saca varillas en doble hasta punto de pesca, completando pozo. Desarmar componentes que requieran reemplazo. "
                    "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                    "Registrar evidencia fotográfica del estado del material y punto de pesca. Asentar en OW grado de acero del material extraído."
                ),               
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""
            })
            
            
             # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm  = row["COMENTARIO"] or ""
                # convierto a float para comparar
                try:
                    dia_val = float(dia)
                except:
                    dia_val = None

                # 1) lógica especial para PESO
                if "PESO" in elem.upper() and dia_val is not None:
                    if abs(dia_val - 1.5) < 1e-3:
                        modo = "(simple)"
                    elif abs(dia_val - 1.125) < 1e-3:
                        modo = "(doble)"
                    else:
                        # cae al modo por defecto
                        modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                else:
                    # 2) modo por comentario
                    modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""

                parts.append(f"{qty} {elem} {dia}in {modo}".strip())

            diseño = " + ".join(parts)

            desc_sv_simple = (
                "Bajar sarta de varillas en doble + pescador y Pescar. Desclavar bomba. "f"Diseño a extraer: {diseño}. "
                "Sacar varillas pescadas, desarmando conexión, completando pozo. "
                "Desarmar componentes que requieran reemplazo. Reemplazar 5 varillas por arriba y por debajo en zona de pesca."
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "Punto programa":        10,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })
            
            # 11) ACONDICIONAMIENTO PARA SACAR CAÑOS
            m.append({ "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA SACAR CAÑOS","Punto programa":11,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, completar con ASDF, desempaquetar y montar conjunto BOP.Montar piso de trabajo.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"252","TIEMPO":"" })

            # 12) SACA TUBING
            m.append({ "MANIOBRA NORMALIZADA":"SACA TUBING","Punto programa":12,
                "DESCRIPCIÓN":"Tareas generales durante la sacada de tubing.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })

            # 13) SACA TUBING EN DOBLE with anchor
            parts_tub = [
                f"{int(r['CANTIDAD'])} {r['ELEMENTO']} {r['DIÁMETRO']}in "
                f"({'DOBLE' if 'DOBLE' in str(r['COMENTARIO']).upper() else 'SIMPLE'})"
                for _,r in tubing_act.iterrows()
            ]
            diseño_tub = " + ".join(parts_tub)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA TUBING EN DOBLE","Punto programa":13,
                "DESCRIPCIÓN":(
                    "Librar ANCLA. Sacar completando pozo. Desarmar componentes que requieran reemplazo. "
                    f"Diseño a extraer: {diseño_tub}. Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído. Solicitar envío de bomba a taller de Pecom de inmediato para desarmar e inspeccionar. Indicar si evidencia falla visible. Asentar número de bomba y estado de cabezal y filtro.."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })    
            
            
            # Maniobra 9: BAJA TUBING EN SIMPLE
            # Construyo “INSTALACIÓN FINAL tubing” a partir de tubing_fin
            
            
            partes_fin_tub = []
            for _, r in tubing_fin.iterrows():
                elem = r["ELEMENTO"]
                if not (isinstance(elem, str) and elem.strip()):
                    continue

                tokens = []
                # cantidad
                if pd.notna(r["CANTIDAD"]):
                    tokens.append(str(int(r["CANTIDAD"])))
                # elemento
                tokens.append(elem.strip())
                # condición
                if pd.notna(r["CONDICIÓN"]):
                    tokens.append(r["CONDICIÓN"].strip())
                # diámetro
                if pd.notna(r["DIÁMETRO"]):
                    tokens.append(f"{r['DIÁMETRO']}in")
                # profundidad
                if pd.notna(r["PROFUNDIDAD"]):
                    tokens.append(f"en {r['PROFUNDIDAD']}m")

                parte = " ".join(tokens)
                # comentario libre
                if pd.notna(r["COMENTARIO"]) and r["COMENTARIO"].strip():
                    parte += f" ({r['COMENTARIO'].strip()})"

                partes_fin_tub.append(parte)
            

            # Invierto el orden
            partes_fin_tub.reverse()

            # Concateno con " + "
            diseño_fin_tub = " + ".join(partes_fin_tub) if partes_fin_tub else ""

            desc9 = (
                "Profundizar columna de TBG calibrando, midiendo, limpiando, engrasando y torqueando las conexiones. "
                "Bajar instalación en simple. "
                f"Diseño a bajar: {diseño_fin_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar número de bomba bajado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "Punto programa":       9,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc9
            })

            # Maniobra 11: ACONDICIONA PARA PH
            desc11 = "Acondicionar superficie para realizar prueba hidráulica."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONA PARA PH",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "259",
                "Punto programa":       11,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc11
            })

            # Maniobra 12: PRUEBA DE HERMETICIDAD
            desc12 = "Realizar PH inicial, intermedia y final con 800, 900 y 1000 psi respectivamente.Registrar la misma en OpenWells.Si la prueba es deficiente informar a Supervisor de Pulling."
            m.append({
                "MANIOBRA NORMALIZADA": "PRUEBA DE HERMETICIDAD",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "205",
                "Punto programa":       12,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc12
            })    
                
            # 20) FIJA ANCLA
            # — Constantes de cálculo de ancla —
            WALL_THICKNESS        = {2.875: 0.434, 3.5: 0.508}
            CONVERSION_PIES_METROS = 3.28084
            COEF_POISSON          = 0.3
            MODULO_YOUNG          = 30_000_000
            COEF_EXPANSION        = 0.0000069
            GRADIENTE_FLUIDO      = 0.5
            TEMP_SUPERFICIE_C     = 30
            TEMP_MEDIA_C          = 12
            NIVEL_ESTATICO_PIE    = 656
            
            # — Validación de PROFUNDIDAD de ANCLA —
            ancla = tubing_fin[tubing_fin["ELEMENTO"].str.contains("ANCLA", case=False, na=False)]
            if ancla.empty or pd.isna(ancla["PROFUNDIDAD"].iloc[0]):
                # Aviso de que faltan datos, pero no abortamos el proceso
                print("❌ Falta la profundidad de ancla. No se calculó tensión ni estiramiento.")
                desc10 = "Falta la profundidad de ancla; no se calculó tensión ni estiramiento."
            else:
                prof_m = float(ancla["PROFUNDIDAD"].iloc[0])

                # 20) FIJA ANCLA
                #diam_ext = float(ancla["DIÁMETRO"].iloc[0])
                diam_ext = 2.875
                thickness = WALL_THICKNESS.get(diam_ext, 0.434)

                prof_ft   = prof_m * CONVERSION_PIES_METROS
                dynamic_m = max(prof_m - (NIVEL_ESTATICO_PIE/CONVERSION_PIES_METROS), 0)
                dynamic_ft= dynamic_m * CONVERSION_PIES_METROS

                thickness = WALL_THICKNESS.get(diam_ext, 0.434)
                diam_int  = diam_ext - thickness

                # ratio para F1
                ratio = ((COEF_POISSON * dynamic_ft / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                F1    = 6.4918 * dynamic_ft * GRADIENTE_FLUIDO * ratio
                seccion_pared = 1.812
                F2    = MODULO_YOUNG * COEF_EXPANSION * ((TEMP_SUPERFICIE_C - TEMP_MEDIA_C)/2) * seccion_pared

                B     = ((COEF_POISSON * NIVEL_ESTATICO_PIE / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                A     = 6.4918 * GRADIENTE_FLUIDO * ((seccion_pared / (diam_ext**2)) * NIVEL_ESTATICO_PIE)
                F3    = A * B

                tension_total   = F1 + F2 - F3
                estiramiento_in = 0.22 * (prof_ft/1000) * (tension_total/1000)
                est_cm          = estiramiento_in * 2.54

                desc10 = f"Fijar ancla con {tension_total:.2f} lbs y {est_cm:.2f} cm de estiramiento."
            m.append({
                "MANIOBRA NORMALIZADA": "FIJA ANCLA",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP17",
                "ACTIVITY_SUBCODE":     "211",
                "Punto programa":       10,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc10
            })    
                
            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo, montar piso de trabajo + herramientas de v/b.."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })

            # Maniobra 14: BAJA VARILLAS
            desc14 = "Tareas varias durante bajada de varillas."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       14,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc14
            })

           # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de bomba y bajarla + sarta de v/b  limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       15,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc15
            })


            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "Punto programa":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "Punto programa":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi.Realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })



            # Maniobra 16: VARIOS
            desc16 = "Tareas varias."
            m.append({
                "MANIOBRA NORMALIZADA": "VARIOS",
                "ACTIVITY_PHASE":       "SPV",
                "ACTIVITY_CODE":        "SPV",
                "ACTIVITY_SUBCODE":     "SPV",
                "Punto programa":       16,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc16
            })

            # Maniobra 17: ARMA BDP
            desc17 = "Arma BDP"
            m.append({
                "MANIOBRA NORMALIZADA": "ARMA BDP",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP15",
                "ACTIVITY_SUBCODE":     "260",
                "Punto programa":       17,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc17
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("generar_bayler_bm_program.xlsx", index=False)
            print("\n✅ Archivo guardado como 'generar_bayler_bm_program.xlsx'")
            
        def generar_recupero_bm():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            
            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo}. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

            # Maniobra 3: DESARMA BDP
            desc3 = "Desarmar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BDP",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP34",
                "ACTIVITY_SUBCODE":     "210",
                "Punto programa":       3,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc3
            })

            # — 6) ACONDICIONA PARA SACAR VARILLAS (con PH inicial)
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","Punto programa":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. Tomar peso de sarta, y registrar en OW.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
                        
            # — 7) SACA VARILLAS (solo peso/anomalías)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS","Punto programa":7,
                "DESCRIPCIÓN":"Tareas varias durante saca de varillas",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""})
            
             # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm  = row["COMENTARIO"] or ""
                # convierto a float para comparar
                try:
                    dia_val = float(dia)
                except:
                    dia_val = None

                # 1) lógica especial para PESO
                if "PESO" in elem.upper() and dia_val is not None:
                    if abs(dia_val - 1.5) < 1e-3:
                        modo = "(simple)"
                    elif abs(dia_val - 1.125) < 1e-3:
                        modo = "(doble)"
                    else:
                        # cae al modo por defecto
                        modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                else:
                    # 2) modo por comentario
                    modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""

                parts.append(f"{qty} {elem} {dia}in {modo}".strip())

            diseño = " + ".join(parts)

            desc_sv_simple = (
                "Desclavar bomba."
                "Sacar sarta desarmando, completando pozo. Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído."
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "Punto programa":        10,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })
            
            # 11) ACONDICIONAMIENTO PARA SACAR CAÑOS
            m.append({ "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA SACAR CAÑOS","Punto programa":11,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, completar con ASDF, desempaquetar y montar conjunto BOP.Montar piso de trabajo.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"252","TIEMPO":"" })

            # 12) SACA TUBING
            m.append({ "MANIOBRA NORMALIZADA":"SACA TUBING","Punto programa":12,
                "DESCRIPCIÓN":"Tareas generales durante la sacada de tubing.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })

            # 13) SACA TUBING EN DOBLE with anchor
            parts_tub = [
                f"{int(r['CANTIDAD'])} {r['ELEMENTO']} {r['DIÁMETRO']}in "
                f"({'DOBLE' if 'DOBLE' in str(r['COMENTARIO']).upper() else 'SIMPLE'})"
                for _,r in tubing_act.iterrows()
            ]
            diseño_tub = " + ".join(parts_tub)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA TUBING EN DOBLE","Punto programa":13,
                "DESCRIPCIÓN":(
                    "Librar ANCLA. Sacar columna en tiro doble, completando pozo. Desarmar componentes que requieran reemplazo. "
                    f"Diseño a extraer: {diseño_tub}. Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído. Solicitar envío de bomba a taller de Pecom de inmediato para desarmar e inspeccionar. Indicar si evidencia falla visible. Asentar número de bomba y estado de cabezal y filtro.."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })    
            
           
            # Maniobra 17: ARMA BDP
            desc17 = "Colocar tapa y valvula."
            m.append({
                "MANIOBRA NORMALIZADA": "ARMA BDP",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP15",
                "ACTIVITY_SUBCODE":     "260",
                "Punto programa":       17,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc17
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación"
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("Recupero_program.xlsx", index=False)
            print("\n✅ Archivo guardado como 'RecuperoBM.xlsx'")
        
        def generar_Sacar_desenrosque_tub_bm():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            
            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo} En intervención anterior no se logró pescar, en pozo se encuentra-----------------------. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

            # Maniobra 3: DESARMA BDP
            desc3 = "Desarmar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BDP",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP34",
                "ACTIVITY_SUBCODE":     "210",
                "Punto programa":       3,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc3
            })

            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo, montar piso de trabajo + herramientas de v/b.."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })
                        
                      
            
             # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm  = row["COMENTARIO"] or ""
                # convierto a float para comparar
                try:
                    dia_val = float(dia)
                except:
                    dia_val = None

                # 1) lógica especial para PESO
                if "PESO" in elem.upper() and dia_val is not None:
                    if abs(dia_val - 1.5) < 1e-3:
                        modo = "(simple)"
                    elif abs(dia_val - 1.125) < 1e-3:
                        modo = "(doble)"
                    else:
                        # cae al modo por defecto
                        modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                else:
                    # 2) modo por comentario
                    modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""

                parts.append(f"{qty} {elem} {dia}in {modo}".strip())

            diseño = " + ".join(parts)

            desc_sv_simple = (
                "Bajar sarta de varillas + pescador y Pescar, ver de lograr desclavar bomba. Si el resultado es negativo, realizar desenrosque y mover cañeria. "f"Diseño a extraer: {diseño_act}. "
                "Indicar punto de desenroque e informar "
                
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "Punto programa":        10,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })
            
            # 11) ACONDICIONAMIENTO PARA SACAR CAÑOS
            m.append({ "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA SACAR CAÑOS","Punto programa":11,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, completar con ASDF, desempaquetar y montar conjunto BOP.Montar piso de trabajo.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"252","TIEMPO":"" })

            # 12) SACA TUBING
            m.append({ "MANIOBRA NORMALIZADA":"SACA TUBING","Punto programa":12,
                "DESCRIPCIÓN":"Tareas generales durante la sacada de tubing.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })

            # 13) SACA TUBING EN DOBLE with anchor
            parts_tub = [
                f"{int(r['CANTIDAD'])} {r['ELEMENTO']} {r['DIÁMETRO']}in "
                f"({'DOBLE' if 'DOBLE' in str(r['COMENTARIO']).upper() else 'SIMPLE'})"
                for _,r in tubing_act.iterrows()
            ]
            diseño_tub = " + ".join(parts_tub)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA TUBING EN DOBLE","Punto programa":13,
                "DESCRIPCIÓN":(
                    "Librar ANCLA. Sacar columna hasta punto de desenrosque, completando pozo.Si se requiera continuar con maniobra de desenrosque de varillas + sacada de tubing hasta punto de pesca. Desarmar componentes que requieran reemplazo. "
                    f"Diseño a extraer: {diseño_tub}. Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído. Solicitar envío de bomba a taller de Pecom de inmediato para desarmar e inspeccionar. Indicar si evidencia falla visible. Asentar número de bomba y estado de cabezal y filtro.."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })    
            
            
            # Maniobra 9: BAJA TUBING EN SIMPLE
            # Construyo “INSTALACIÓN FINAL tubing” a partir de tubing_fin
            
            
            partes_fin_tub = []
            for _, r in tubing_fin.iterrows():
                elem = r["ELEMENTO"]
                if not (isinstance(elem, str) and elem.strip()):
                    continue

                tokens = []
                # cantidad
                if pd.notna(r["CANTIDAD"]):
                    tokens.append(str(int(r["CANTIDAD"])))
                # elemento
                tokens.append(elem.strip())
                # condición
                if pd.notna(r["CONDICIÓN"]):
                    tokens.append(r["CONDICIÓN"].strip())
                # diámetro
                if pd.notna(r["DIÁMETRO"]):
                    tokens.append(f"{r['DIÁMETRO']}in")
                # profundidad
                if pd.notna(r["PROFUNDIDAD"]):
                    tokens.append(f"en {r['PROFUNDIDAD']}m")

                parte = " ".join(tokens)
                # comentario libre
                if pd.notna(r["COMENTARIO"]) and r["COMENTARIO"].strip():
                    parte += f" ({r['COMENTARIO'].strip()})"

                partes_fin_tub.append(parte)
            

            # Invierto el orden
            partes_fin_tub.reverse()

            # Concateno con " + "
            diseño_fin_tub = " + ".join(partes_fin_tub) if partes_fin_tub else ""

            desc9 = (
                "Profundizar columna de TBG calibrando, midiendo, limpiando, engrasando y torqueando las conexiones. "
                "Bajar instalación en simple. "
                f"Diseño a bajar: {diseño_fin_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar número de bomba bajado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "Punto programa":       9,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc9
            })

            # Maniobra 11: ACONDICIONA PARA PH
            desc11 = "Acondicionar superficie para realizar prueba hidráulica."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONA PARA PH",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "259",
                "Punto programa":       11,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc11
            })

            # Maniobra 12: PRUEBA DE HERMETICIDAD
            desc12 = "Realizar PH inicial, intermedia y final con 800, 900 y 1000 psi respectivamente.Registrar la misma en OpenWells.Si la prueba es deficiente informar a Supervisor de Pulling."
            m.append({
                "MANIOBRA NORMALIZADA": "PRUEBA DE HERMETICIDAD",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "205",
                "Punto programa":       12,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc12
            })    
                
            # 20) FIJA ANCLA
            # — Constantes de cálculo de ancla —
            WALL_THICKNESS        = {2.875: 0.434, 3.5: 0.508}
            CONVERSION_PIES_METROS = 3.28084
            COEF_POISSON          = 0.3
            MODULO_YOUNG          = 30_000_000
            COEF_EXPANSION        = 0.0000069
            GRADIENTE_FLUIDO      = 0.5
            TEMP_SUPERFICIE_C     = 30
            TEMP_MEDIA_C          = 12
            NIVEL_ESTATICO_PIE    = 656
            
            # — Validación de PROFUNDIDAD de ANCLA —
            ancla = tubing_fin[tubing_fin["ELEMENTO"].str.contains("ANCLA", case=False, na=False)]
            if ancla.empty or pd.isna(ancla["PROFUNDIDAD"].iloc[0]):
                # Aviso de que faltan datos, pero no abortamos el proceso
                print("❌ Falta la profundidad de ancla. No se calculó tensión ni estiramiento.")
                desc10 = "Falta la profundidad de ancla; no se calculó tensión ni estiramiento."
            else:
                prof_m = float(ancla["PROFUNDIDAD"].iloc[0])

                # 20) FIJA ANCLA
                #diam_ext = float(ancla["DIÁMETRO"].iloc[0])
                diam_ext = 2.875
                thickness = WALL_THICKNESS.get(diam_ext, 0.434)

                prof_ft   = prof_m * CONVERSION_PIES_METROS
                dynamic_m = max(prof_m - (NIVEL_ESTATICO_PIE/CONVERSION_PIES_METROS), 0)
                dynamic_ft= dynamic_m * CONVERSION_PIES_METROS

                thickness = WALL_THICKNESS.get(diam_ext, 0.434)
                diam_int  = diam_ext - thickness

                # ratio para F1
                ratio = ((COEF_POISSON * dynamic_ft / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                F1    = 6.4918 * dynamic_ft * GRADIENTE_FLUIDO * ratio
                seccion_pared = 1.812
                F2    = MODULO_YOUNG * COEF_EXPANSION * ((TEMP_SUPERFICIE_C - TEMP_MEDIA_C)/2) * seccion_pared

                B     = ((COEF_POISSON * NIVEL_ESTATICO_PIE / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                A     = 6.4918 * GRADIENTE_FLUIDO * ((seccion_pared / (diam_ext**2)) * NIVEL_ESTATICO_PIE)
                F3    = A * B

                tension_total   = F1 + F2 - F3
                estiramiento_in = 0.22 * (prof_ft/1000) * (tension_total/1000)
                est_cm          = estiramiento_in * 2.54

                desc10 = f"Fijar ancla con {tension_total:.2f} lbs y {est_cm:.2f} cm de estiramiento."
            m.append({
                "MANIOBRA NORMALIZADA": "FIJA ANCLA",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP17",
                "ACTIVITY_SUBCODE":     "211",
                "Punto programa":       10,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc10
            })    
                
            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo, montar piso de trabajo + herramientas de v/b.."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })

            # Maniobra 14: BAJA VARILLAS
            desc14 = "Tareas varias durante bajada de varillas."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       14,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc14
            })

           # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de bomba y bajarla + sarta de v/b  limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       15,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc15
            })


            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "Punto programa":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "Punto programa":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi.Realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })



            # Maniobra 16: VARIOS
            desc16 = "Tareas varias."
            m.append({
                "MANIOBRA NORMALIZADA": "VARIOS",
                "ACTIVITY_PHASE":       "SPV",
                "ACTIVITY_CODE":        "SPV",
                "ACTIVITY_SUBCODE":     "SPV",
                "Punto programa":       16,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc16
            })

            # Maniobra 17: ARMA BDP
            desc17 = "Arma BDP"
            m.append({
                "MANIOBRA NORMALIZADA": "ARMA BDP",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP15",
                "ACTIVITY_SUBCODE":     "260",
                "Punto programa":       17,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc17
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("generar_parcialvarillas_movtbg_program.xlsx", index=False)
            print("\n✅ Archivo guardado como 'generar_parcialvarillas_movtbg_program.xlsx'")
        
        
        
        def generar_cambio_PCP_x_bm():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            # — Constantes de cálculo de ancla —
            WALL_THICKNESS        = {2.875: 0.434, 3.5: 0.508}
            CONVERSION_PIES_METROS = 3.28084
            COEF_POISSON          = 0.3
            MODULO_YOUNG          = 30_000_000
            COEF_EXPANSION        = 0.0000069
            GRADIENTE_FLUIDO      = 0.5
            TEMP_SUPERFICIE_C     = 30
            TEMP_MEDIA_C          = 12
            NIVEL_ESTATICO_PIE    = 656

            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo}. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

            # Maniobra 3: DESARMA BDP
            desc3 = "Desarmar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BDP",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP34",
                "ACTIVITY_SUBCODE":     "210",
                "Punto programa":       3,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc3
            })
            
            
             # — 6) ACONDICIONA PARA SACAR VARILLAS (con PH inicial)
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","Punto programa":6,
                "DESCRIPCIÓN":"Desmontar cabezal de PCP conjunto con compañía SLB, verificar giro de rotor, salida y entrada  y librar rotor.Acondicionar boca de pozo, montar piso de trabajo + htas de v/b.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
                        
            # — 7) SACA VARILLAS (solo peso/anomalías)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS","Punto programa":7,
                "DESCRIPCIÓN":"Tareas varias durante saca de varillas",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""})
            
             # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                parts.append(f"{qty} {elem} {dia}in {modo}".strip())
            diseño = " + ".join(parts)

            desc_sv_simple = (
                ". Sacar sarta, desarmando, completando pozo. Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído."
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "Punto programa":        10,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })
            
            # Maniobra 4: ACONDICIONAMIENTO PARA SACAR CAÑOS
            desc4 = (
                "Acondicionar boca de pozo, completar con ASDF, desempaquetar y montar conjunto BOP anular y spooler. "
                "Montar piso de trabajo."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA SACAR CAÑOS",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP24",
                "ACTIVITY_SUBCODE":     "252",
                "Punto programa":       4,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc4
            })
            
            
            # 12) SACA TUBING
            m.append({ "MANIOBRA NORMALIZADA":"SACA TUBING","Punto programa":12,
                "DESCRIPCIÓN":"Tareas generales durante la sacada de tubing.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })

            # 13) SACA TUBING EN DOBLE with anchor
            
            # Construyo “INSTALACIÓN ACTUAL tubing” a partir de tubing_act
            partes_act_tub = []
            for _, r in tubing_act.iterrows():
                if isinstance(r["ELEMENTO"], str) and r["ELEMENTO"].strip():
                    # Cantidad y diámetro siempre presentes
                    cant = int(r["CANTIDAD"]) if pd.notna(r["CANTIDAD"]) else 0
                    diam = r["DIÁMETRO"]
                    elem = r["ELEMENTO"]
                    # Incluyo COMENTARIO si no es NaN ni cadena vacía
                    comm = r["COMENTARIO"] if (isinstance(r["COMENTARIO"], str) and r["COMENTARIO"].strip()) else ""
                    if comm:
                        partes_act_tub.append(f"{cant} {elem} {diam}in {comm}")
                    else:
                        partes_act_tub.append(f"{cant} {elem} {diam}in")
            # Armo el string completo, separando con “ + ”
            diseño_act_tub = " + ".join(partes_act_tub) if partes_act_tub else ""

            desc5 = (
                "Desplazar por directa para sacar sarta limpia. Sacar sarta, buscando pérdida y completando pozo. "
                "Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño_act_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído. "
                "Solicitar envío de bomba PC a taller para desarmar e inspeccionar. Indicar si evidencia falla visible. "
                "Asentar número de bomba PCP y estado de componentes."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "SACA TUBING EN DOBLE",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP24",
                "ACTIVITY_SUBCODE":     "253",
                "Punto programa":       5,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc5
            })

            ## Maniobra 7: ACONDICIONAMIENTO PARA BAJAR CAÑOS
            ##esc7 = "Acondicionar boca de pozo, completar con ASDF, desempaquetar y montar conjunto BOP anular."
            #m#.append({
             #   "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR CAÑOS",
              #  "ACTIVITY_PHASE":       "SP04",
               ## "ACTIVITY_CODE":        "SP16",
              #  "ACTIVITY_SUBCODE":     "256",
              #  "Punto programa":       7,
              #  "TIEMPO":               "",
              #  "DESCRIPCIÓN":          desc7
         #   })

            # Maniobra 8: BAJA TUBING
            desc8 = "Tareas generales durante la bajada de tubing."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "Punto programa":       8,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc8
            })
            
            # Maniobra 9: BAJA TUBING EN SIMPLE
            # Construyo “INSTALACIÓN FINAL tubing” a partir de tubing_fin
            
            
            partes_fin_tub = []
            for _, r in tubing_fin.iterrows():
                elem = r["ELEMENTO"]
                if not (isinstance(elem, str) and elem.strip()):
                    continue

                tokens = []
                # cantidad
                if pd.notna(r["CANTIDAD"]):
                    tokens.append(str(int(r["CANTIDAD"])))
                # elemento
                tokens.append(elem.strip())
                # condición
                if pd.notna(r["CONDICIÓN"]):
                    tokens.append(r["CONDICIÓN"].strip())
                # diámetro
                if pd.notna(r["DIÁMETRO"]):
                    tokens.append(f"{r['DIÁMETRO']}in")
                # profundidad
                if pd.notna(r["PROFUNDIDAD"]):
                    tokens.append(f"en {r['PROFUNDIDAD']}m")

                parte = " ".join(tokens)
                # comentario libre
                if pd.notna(r["COMENTARIO"]) and r["COMENTARIO"].strip():
                    parte += f" ({r['COMENTARIO'].strip()})"

                partes_fin_tub.append(parte)
            

            # Invierto el orden
            partes_fin_tub.reverse()

            # Concateno con " + "
            diseño_fin_tub = " + ".join(partes_fin_tub) if partes_fin_tub else ""

            desc9 = (
                "Profundizar columna de TBG calibrando, midiendo, limpiando, engrasando y torqueando las conexiones. "
                "Bajar instalación en simple. "
                f"Diseño a bajar: {diseño_fin_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar número de bomba bajado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "Punto programa":       9,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc9
            })

            # Maniobra 11: ACONDICIONA PARA PH
            desc11 = "Acondicionar superficie para realizar prueba hidráulica."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONA PARA PH",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "259",
                "Punto programa":       11,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc11
            })

            # Maniobra 12: PRUEBA DE HERMETICIDAD
            desc12 = "Realizar PH inicial, intermedia y final con 800, 900 y 1000 psi respectivamente. Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling. . Si la prueba es deficiente informar a Supervisor de Pulling."
            m.append({
                "MANIOBRA NORMALIZADA": "PRUEBA DE HERMETICIDAD",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "205",
                "Punto programa":       12,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc12
            })


            ancla = tubing_fin[tubing_fin["ELEMENTO"].str.contains("ANCLA", case=False, na=False)]
            if ancla.empty or pd.isna(ancla["PROFUNDIDAD"].iloc[0]):
                # Aviso de que faltan datos, pero no abortamos el proceso
                print("❌ Falta la profundidad de ancla. No se calculó tensión ni estiramiento.")
                desc10 = "Falta la profundidad de ancla; no se calculó tensión ni estiramiento."
            else:
                prof_m = float(ancla["PROFUNDIDAD"].iloc[0])

                # 20) FIJA ANCLA
                #diam_ext = float(ancla["DIÁMETRO"].iloc[0])
                diam_ext = 2.875
                thickness = WALL_THICKNESS.get(diam_ext, 0.434)

                prof_ft   = prof_m * CONVERSION_PIES_METROS
                dynamic_m = max(prof_m - (NIVEL_ESTATICO_PIE/CONVERSION_PIES_METROS), 0)
                dynamic_ft= dynamic_m * CONVERSION_PIES_METROS

                thickness = WALL_THICKNESS.get(diam_ext, 0.434)
                diam_int  = diam_ext - thickness

                # ratio para F1
                ratio = ((COEF_POISSON * dynamic_ft / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                F1    = 6.4918 * dynamic_ft * GRADIENTE_FLUIDO * ratio
                seccion_pared = 1.812
                F2    = MODULO_YOUNG * COEF_EXPANSION * ((TEMP_SUPERFICIE_C - TEMP_MEDIA_C)/2) * seccion_pared

                B     = ((COEF_POISSON * NIVEL_ESTATICO_PIE / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                A     = 6.4918 * GRADIENTE_FLUIDO * ((seccion_pared / (diam_ext**2)) * NIVEL_ESTATICO_PIE)
                F3    = A * B

                tension_total   = F1 + F2 - F3
                estiramiento_in = 0.22 * (prof_ft/1000) * (tension_total/1000)
                est_cm          = estiramiento_in * 2.54
            
                
                desc10 = f"Fijar ancla con {tension_total:.2f} lbs y {est_cm:.2f} cm de estiramiento."
            m.append({
                "MANIOBRA NORMALIZADA": "FIJA ANCLA",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP17",
                "ACTIVITY_SUBCODE":     "211",
                "Punto programa":       10,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc10
            })


            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo, montar piso de trabajo + herramientas de v/b."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })

            # Maniobra 14: BAJA VARILLAS
            desc14 = "Tareas varias durante bajada de varillas."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       14,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc14
            })

           # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de rotor y bajarlo + sarta de v/b en simple limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. Realizar medida correctamente "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       15,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc15
            })


            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "Punto programa":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "Punto programa":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre rotor con 1000 psi."
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })



            # Maniobra 16: VARIOS
            desc16 = "Tareas varias."
            m.append({
                "MANIOBRA NORMALIZADA": "VARIOS",
                "ACTIVITY_PHASE":       "SPV",
                "ACTIVITY_CODE":        "SPV",
                "ACTIVITY_SUBCODE":     "SPV",
                "Punto programa":       16,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc16
            })

            # Maniobra 17: ARMA BDP
            desc17 = "Arma BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "ARMA BDP",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP15",
                "ACTIVITY_SUBCODE":     "260",
                "Punto programa":       17,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc17
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("PCPXBM_program.xlsx", index=False)
        
        
        def generar_probador_asentado():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            
            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo}. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

            # Maniobra 3: DESARMA BDP
            desc3 = "Desarmar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BDP",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP34",
                "ACTIVITY_SUBCODE":     "210",
                "Punto programa":       3,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc3
            })

            # — 6) ACONDICIONA PARA SACAR VARILLAS (con PH inicial)
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA SACAR VARILLAS","Punto programa":6,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, montar piso de trabajo + htas de v/b. Retirar vástago completo. Tomar peso de sarta, y registrar en OW. Realizar PH inicial con 1000 psi e informar resultados, de ser positiva continuar con maniobras.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"250","TIEMPO":""
            })
                        
            # — 7) SACA VARILLAS (solo peso/anomalías)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA VARILLAS","Punto programa":7,
                "DESCRIPCIÓN":"Registrar peso de la sarta y eventuales anomalías antes de iniciar la extracción.Tareas varias durante saca de varillas",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"251","TIEMPO":""})
            
             # —10) SACA VARILLAS EN SIMPLE
            # Construyo el diseño a extraer a partir de varillas_act (Instalación Actual Varillas)
            parts = []
            for _, row in varillas_act.iterrows():
                qty  = int(row["CANTIDAD"])
                elem = row["ELEMENTO"]
                dia  = row["DIÁMETRO"]
                comm  = row["COMENTARIO"] or ""
                # convierto a float para comparar
                try:
                    dia_val = float(dia)
                except:
                    dia_val = None

                # 1) lógica especial para PESO
                if "PESO" in elem.upper() and dia_val is not None:
                    if abs(dia_val - 1.5) < 1e-3:
                        modo = "(simple)"
                    elif abs(dia_val - 1.125) < 1e-3:
                        modo = "(doble)"
                    else:
                        # cae al modo por defecto
                        modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""
                else:
                    # 2) modo por comentario
                    modo = "(simple)" if "SIMPLE" in comm.upper() else "(doble)" if "DOBLE" in comm.upper() else ""

                parts.append(f"{qty} {elem} {dia}in {modo}".strip())

            diseño = " + ".join(parts)

            desc_sv_simple = (
                "Desclavar bomba. "
                "Sacar sarta desarmando, completando pozo. Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído."
            )
            m.append({
                "MANIOBRA NORMALIZADA":  "SACA VARILLAS EN SIMPLE",
                "Punto programa":        10,
                "DESCRIPCIÓN":           desc_sv_simple,
                "ACTIVITY_PHASE":        "SP03",
                "ACTIVITY_CODE":         "SP24",
                "ACTIVITY_SUBCODE":      "251",
                "TIEMPO":                ""
            })
            
            # 11) ACONDICIONAMIENTO PARA SACAR CAÑOS
            m.append({ "MANIOBRA NORMALIZADA":"ACONDICIONAMIENTO PARA SACAR CAÑOS","Punto programa":11,
                "DESCRIPCIÓN":"Acondicionar boca de pozo, completar con ASDF, desempaquetar y montar conjunto BOP.Montar piso de trabajo.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"252","TIEMPO":"" })

            # 12) SACA TUBING
            m.append({ "MANIOBRA NORMALIZADA":"SACA TUBING","Punto programa":12,
                "DESCRIPCIÓN":"Tareas generales durante la sacada de tubing.","ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })

            # 13) SACA TUBING EN DOBLE with anchor
            parts_tub = [
                f"{int(r['CANTIDAD'])} {r['ELEMENTO']} {r['DIÁMETRO']}in "
                f"({'DOBLE' if 'DOBLE' in str(r['COMENTARIO']).upper() else 'SIMPLE'})"
                for _,r in tubing_act.iterrows()
            ]
            diseño_tub = " + ".join(parts_tub)
            m.append({
                "MANIOBRA NORMALIZADA":"SACA TUBING DESAGOTANDO","Punto programa":13,
                "DESCRIPCIÓN":(
                    "Librar ANCLA. Bajar probador, llenar columna con ASDF,Sacar buscando pérdida.Reemplazar 10 tubing por encima y por debajo de pérdida. Desarmar componentes que requieran reemplazo. "
                    f"Diseño a extraer: {diseño_tub}. Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído. Solicitar envío de bomba a taller de Pecom de inmediato para desarmar e inspeccionar. Indicar si evidencia falla visible. Asentar número de bomba y estado de cabezal y filtro.."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP24","ACTIVITY_SUBCODE":"253","TIEMPO":"" })    
            
            
                       
            # Maniobra 9: BAJA TUBING EN SIMPLE
            # Construyo “INSTALACIÓN FINAL tubing” a partir de tubing_fin
            
            
            partes_fin_tub = []
            for _, r in tubing_fin.iterrows():
                elem = r["ELEMENTO"]
                if not (isinstance(elem, str) and elem.strip()):
                    continue

                tokens = []
                # cantidad
                if pd.notna(r["CANTIDAD"]):
                    tokens.append(str(int(r["CANTIDAD"])))
                # elemento
                tokens.append(elem.strip())
                # condición
                if pd.notna(r["CONDICIÓN"]):
                    tokens.append(r["CONDICIÓN"].strip())
                # diámetro
                if pd.notna(r["DIÁMETRO"]):
                    tokens.append(f"{r['DIÁMETRO']}in")
                # profundidad
                if pd.notna(r["PROFUNDIDAD"]):
                    tokens.append(f"en {r['PROFUNDIDAD']}m")

                parte = " ".join(tokens)
                # comentario libre
                if pd.notna(r["COMENTARIO"]) and r["COMENTARIO"].strip():
                    parte += f" ({r['COMENTARIO'].strip()})"

                partes_fin_tub.append(parte)
            

            # Invierto el orden
            partes_fin_tub.reverse()

            # Concateno con " + "
            diseño_fin_tub = " + ".join(partes_fin_tub) if partes_fin_tub else ""

            desc9 = (
                "Profundizar columna de TBG calibrando, midiendo, limpiando, engrasando y torqueando las conexiones. "
                "Bajar instalación. "
                f"Diseño a bajar: {diseño_fin_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar número de bomba bajado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "Punto programa":       9,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc9
            })

            # Maniobra 11: ACONDICIONA PARA PH
            desc11 = "Acondicionar superficie para realizar prueba hidráulica."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONA PARA PH",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "259",
                "Punto programa":       11,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc11
            })

            # Maniobra 12: PRUEBA DE HERMETICIDAD
            desc12 = "Realizar PH inicial, intermedia y final con 800, 900 y 1000 psi respectivamente.Registrar la misma en OpenWells.. Librar SO con 2000 psi.Si la prueba es deficiente informar a Supervisor de Pulling."
            m.append({
                "MANIOBRA NORMALIZADA": "PRUEBA DE HERMETICIDAD",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "205",
                "Punto programa":       12,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc12
            })    
                
            # 20) FIJA ANCLA
            # — Constantes de cálculo de ancla —
            WALL_THICKNESS        = {2.875: 0.434, 3.5: 0.508}
            CONVERSION_PIES_METROS = 3.28084
            COEF_POISSON          = 0.3
            MODULO_YOUNG          = 30_000_000
            COEF_EXPANSION        = 0.0000069
            GRADIENTE_FLUIDO      = 0.5
            TEMP_SUPERFICIE_C     = 30
            TEMP_MEDIA_C          = 12
            NIVEL_ESTATICO_PIE    = 656
            
            # — Validación de PROFUNDIDAD de ANCLA —
            ancla = tubing_fin[tubing_fin["ELEMENTO"].str.contains("ANCLA", case=False, na=False)]
            if ancla.empty or pd.isna(ancla["PROFUNDIDAD"].iloc[0]):
                # Aviso de que faltan datos, pero no abortamos el proceso
                print("❌ Falta la profundidad de ancla. No se calculó tensión ni estiramiento.")
                desc10 = "Falta la profundidad de ancla; no se calculó tensión ni estiramiento."
            else:
                prof_m = float(ancla["PROFUNDIDAD"].iloc[0])

                # 20) FIJA ANCLA
                diam_ext = float(ancla["DIÁMETRO"].iloc[0])
                thickness = WALL_THICKNESS.get(diam_ext, 0.434)

                prof_ft   = prof_m * CONVERSION_PIES_METROS
                dynamic_m = max(prof_m - (NIVEL_ESTATICO_PIE/CONVERSION_PIES_METROS), 0)
                dynamic_ft= dynamic_m * CONVERSION_PIES_METROS

                thickness = WALL_THICKNESS.get(diam_ext, 0.434)
                diam_int  = diam_ext - thickness

                # ratio para F1
                ratio = ((COEF_POISSON * dynamic_ft / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                F1    = 6.4918 * dynamic_ft * GRADIENTE_FLUIDO * ratio
                seccion_pared = 1.812
                F2    = MODULO_YOUNG * COEF_EXPANSION * ((TEMP_SUPERFICIE_C - TEMP_MEDIA_C)/2) * seccion_pared

                B     = ((COEF_POISSON * NIVEL_ESTATICO_PIE / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                A     = 6.4918 * GRADIENTE_FLUIDO * ((seccion_pared / (diam_ext**2)) * NIVEL_ESTATICO_PIE)
                F3    = A * B

                tension_total   = F1 + F2 - F3
                estiramiento_in = 0.22 * (prof_ft/1000) * (tension_total/1000)
                est_cm          = estiramiento_in * 2.54

                desc10 = f"Fijar ancla con {tension_total:.2f} lbs y {est_cm:.2f} cm de estiramiento."
            m.append({
                "MANIOBRA NORMALIZADA": "FIJA ANCLA",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP17",
                "ACTIVITY_SUBCODE":     "211",
                "Punto programa":       10,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc10
            })    
                
            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo, montar piso de trabajo + herramientas de v/b.."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })

            # Maniobra 14: BAJA VARILLAS
            desc14 = "Tareas varias durante bajada de varillas."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       14,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc14
            })

           # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de bomba y bajarla + sarta de v/b  limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       15,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc15
            })


            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "Punto programa":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "Punto programa":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi.Realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })



            # Maniobra 16: VARIOS
            desc16 = "Tareas varias."
            m.append({
                "MANIOBRA NORMALIZADA": "VARIOS",
                "ACTIVITY_PHASE":       "SPV",
                "ACTIVITY_CODE":        "SPV",
                "ACTIVITY_SUBCODE":     "SPV",
                "Punto programa":       16,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc16
            })

            # Maniobra 17: ARMA BDP
            desc17 = "Realizar final sobre bomba con 1000 psi. Realizar prueba de bomba. Si la prueba es deficiente informar a Supervisor de Pulling."
            m.append({
                "MANIOBRA NORMALIZADA": "ARMA BDP",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP15",
                "ACTIVITY_SUBCODE":     "260",
                "Punto programa":       17,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc17
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("generar_probador_asentado_program.xlsx", index=False)
            print("\n✅ Archivo guardado como 'probadorasentado_bm_program.xlsx'")
        
        
        
        def generar_cambio_bes_x_bm():
            import math
            import pandas as pd

            global varillas_act, tubing_act, varillas_fin, tubing_fin

            # — Constantes de cálculo de ancla —
            WALL_THICKNESS        = {2.875: 0.434, 3.5: 0.508}
            CONVERSION_PIES_METROS = 3.28084
            COEF_POISSON          = 0.3
            MODULO_YOUNG          = 30_000_000
            COEF_EXPANSION        = 0.0000069
            GRADIENTE_FLUIDO      = 0.5
            TEMP_SUPERFICIE_C     = 30
            TEMP_MEDIA_C          = 12
            NIVEL_ESTATICO_PIE    = 656

            # 0) Inputs desde el datasheet
            meta = {k: df0.iat[r, c] for k, (r, c) in meta_pos.items()}
            pozo        = meta["POZO"]
            definicion  = meta["DEFINICIÓN"]
            motivo      = meta["MANIOBRAS (MOTIVO)"]
            rig_name    = meta["EQUIPO"]

            # concatenar antecedentes no vacíos
            antecedentes = ", ".join(
                [meta[k] for k in ["ANTECEDENTE_1","ANTECEDENTE_2","ANTECEDENTE_3","ANTECEDENTE_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )
            # concatenar requerimientos especiales no vacíos
            reqs = ", ".join(
                [meta[k] for k in ["REQ_ESP_1","REQ_ESP_2","REQ_ESP_3","REQ_ESP_4"]
                 if isinstance(meta.get(k), str) and meta.get(k).strip()]
            )

            # 1) Construyo lista 'm' empezando vacía y luego agrego todas las maniobras según el orden que pediste.
            m = []

            # Maniobra 1: EQUIPO EN TRANSPORTE
            desc1 = (
                f"Transportar a {pozo}. Tener en cuenta los antecedentes del pozo: {antecedentes}. "
                f"La definición actual del pozo es: {definicion}. La maniobra a realizar es {motivo}. "
                f"Considerar los siguientes requerimientos: {reqs}."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "EQUIPO EN TRANSPORTE",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "200",
                "Punto programa":       1,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc1
            })

            # Maniobra 2: MONTAJE EQUIPO
            desc2 = (
                "Verificar presiones por directa y por entrecaño. Desarmar puente de producción. "
                "Montar equipo según procedimiento."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "MONTAJE EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP10",
                "ACTIVITY_SUBCODE":     "201",
                "Punto programa":       2,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc2
            })

            # Maniobra 3: DESARMA BDP
            desc3 = "Desarmar BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BDP",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP34",
                "ACTIVITY_SUBCODE":     "210",
                "Punto programa":       3,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc3
            })

            # Maniobra 4: ACONDICIONAMIENTO PARA SACAR CAÑOS
            desc4 = (
                "Acondicionar boca de pozo, completar con ASDF, desempaquetar y montar conjunto BOP anular y spooler. "
                "Montar piso de trabajo."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA SACAR CAÑOS",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP24",
                "ACTIVITY_SUBCODE":     "252",
                "Punto programa":       4,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc4
            })

                # Maniobra 5: SACA TUBING BES
            # Construyo “INSTALACIÓN ACTUAL tubing” a partir de tubing_act
            partes_act_tub = []
            for _, r in tubing_act.iterrows():
                if isinstance(r["ELEMENTO"], str) and r["ELEMENTO"].strip():
                    # Cantidad y diámetro siempre presentes
                    cant = int(r["CANTIDAD"]) if pd.notna(r["CANTIDAD"]) else 0
                    diam = r["DIÁMETRO"]
                    elem = r["ELEMENTO"]
                    # Incluyo COMENTARIO si no es NaN ni cadena vacía
                    comm = r["COMENTARIO"] if (isinstance(r["COMENTARIO"], str) and r["COMENTARIO"].strip()) else ""
                    if comm:
                        partes_act_tub.append(f"{cant} {elem} {diam}in {comm}")
                    else:
                        partes_act_tub.append(f"{cant} {elem} {diam}in")
            # Armo el string completo, separando con “ + ”
            diseño_act_tub = " + ".join(partes_act_tub) if partes_act_tub else ""

            desc5 = (
                "Sacar columna, sacando cable, buscando pérdida y completando pozo. "
                "Desarmar componentes que requieran reemplazo. "
                f"Diseño a extraer: {diseño_act_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar en OW grado de acero del material extraído. "
                "Solicitar envío de bomba a taller de inmediato para desarmar e inspeccionar. Indicar si evidencia falla visible. "
                "Asentar número de bomba y estado de componentes."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "SACA TUBING BES",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP24",
                "ACTIVITY_SUBCODE":     "253",
                "Punto programa":       5,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc5
            })

            # Maniobra 6: DESARMA BES
            desc6 = "Desarmar conjunto BES."
            m.append({
                "MANIOBRA NORMALIZADA": "DESARMA BES",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP25",
                "ACTIVITY_SUBCODE":     "229",
                "Punto programa":       6,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc6
            })

            # Maniobra 7: ACONDICIONAMIENTO PARA BAJAR CAÑOS
            desc7 = "Acondicionar boca de pozo, completar con ASDF."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR CAÑOS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "256",
                "Punto programa":       7,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc7
            })

            # Maniobra 8: BAJA TUBING
            desc8 = "Tareas generales durante la bajada de tubing."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "Punto programa":       8,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc8
            })

            # Maniobra 9: BAJA TUBING EN SIMPLE
            # Construyo “INSTALACIÓN FINAL tubing” a partir de tubing_fin
            
            
            partes_fin_tub = []
            for _, r in tubing_fin.iterrows():
                elem = r["ELEMENTO"]
                if not (isinstance(elem, str) and elem.strip()):
                    continue

                tokens = []
                # cantidad
                if pd.notna(r["CANTIDAD"]):
                    tokens.append(str(int(r["CANTIDAD"])))
                # elemento
                tokens.append(elem.strip())
                # condición
                if pd.notna(r["CONDICIÓN"]):
                    tokens.append(r["CONDICIÓN"].strip())
                # diámetro
                if pd.notna(r["DIÁMETRO"]):
                    tokens.append(f"{r['DIÁMETRO']}in")
                # profundidad
                if pd.notna(r["PROFUNDIDAD"]):
                    tokens.append(f"en {r['PROFUNDIDAD']}m")

                parte = " ".join(tokens)
                # comentario libre
                if pd.notna(r["COMENTARIO"]) and r["COMENTARIO"].strip():
                    parte += f" ({r['COMENTARIO'].strip()})"

                partes_fin_tub.append(parte)
            
            
            

            # Invierto el orden
            partes_fin_tub.reverse()

            # Concateno con " + "
            diseño_fin_tub = " + ".join(partes_fin_tub) if partes_fin_tub else ""

            desc9 = (
                "Profundizar columna de TBG calibrando, midiendo, limpiando, engrasando y torqueando las conexiones. "
                "Bajar instalación en simple. "
                f"Diseño a bajar: {diseño_fin_tub}. "
                "Asentar en OW observaciones significativas en cuanto a eventual presencia de corrosión, desgaste, incrustaciones o sobretorque. "
                "Registrar evidencia fotográfica del estado del material. Asentar número de bomba bajado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA TUBING EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "257",
                "Punto programa":       9,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc9
            })

            # Maniobra 11: ACONDICIONA PARA PH
            desc11 = "Acondicionar superficie para realizar prueba hidráulica."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONA PARA PH",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "259",
                "Punto programa":       11,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc11
            })

            # Maniobra 12: PRUEBA DE HERMETICIDAD
            desc12 = "Realizar final sobre con 1000 psi. Si la prueba es deficiente informar a Supervisor de Pulling."
            m.append({
                "MANIOBRA NORMALIZADA": "PRUEBA DE HERMETICIDAD",
                "ACTIVITY_PHASE":       "SP03",
                "ACTIVITY_CODE":        "SP13",
                "ACTIVITY_SUBCODE":     "205",
                "Punto programa":       12,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc12
            })


            ancla = tubing_fin[tubing_fin["ELEMENTO"].str.contains("ANCLA", case=False, na=False)]
            if ancla.empty or pd.isna(ancla["PROFUNDIDAD"].iloc[0]):
                # Aviso de que faltan datos, pero no abortamos el proceso
                print("❌ Falta la profundidad de ancla. No se calculó tensión ni estiramiento.")
                desc10 = "Falta la profundidad de ancla; no se calculó tensión ni estiramiento."
            else:
                prof_m = float(ancla["PROFUNDIDAD"].iloc[0])

                # 20) FIJA ANCLA
                diam_ext = float(ancla["DIÁMETRO"].iloc[0])
                thickness = WALL_THICKNESS.get(diam_ext, 0.434)

                prof_ft   = prof_m * CONVERSION_PIES_METROS
                dynamic_m = max(prof_m - (NIVEL_ESTATICO_PIE/CONVERSION_PIES_METROS), 0)
                dynamic_ft= dynamic_m * CONVERSION_PIES_METROS

                thickness = WALL_THICKNESS.get(diam_ext, 0.434)
                diam_int  = diam_ext - thickness

                # ratio para F1
                ratio = ((COEF_POISSON * dynamic_ft / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                F1    = 6.4918 * dynamic_ft * GRADIENTE_FLUIDO * ratio
                seccion_pared = 1.812
                F2    = MODULO_YOUNG * COEF_EXPANSION * ((TEMP_SUPERFICIE_C - TEMP_MEDIA_C)/2) * seccion_pared

                B     = ((COEF_POISSON * NIVEL_ESTATICO_PIE / prof_ft) + (1 - 2 * COEF_POISSON)) if prof_ft > 0 else (1 - 2 * COEF_POISSON)
                A     = 6.4918 * GRADIENTE_FLUIDO * ((seccion_pared / (diam_ext**2)) * NIVEL_ESTATICO_PIE)
                F3    = A * B

                tension_total   = F1 + F2 - F3
                estiramiento_in = 0.22 * (prof_ft/1000) * (tension_total/1000)
                est_cm          = estiramiento_in * 2.54
            
                
                desc10 = f"Fijar ancla con {tension_total:.2f} lbs y {est_cm:.2f} cm de estiramiento."
            m.append({
                "MANIOBRA NORMALIZADA": "FIJA ANCLA",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP17",
                "ACTIVITY_SUBCODE":     "211",
                "Punto programa":       10,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc10
            })


            # Maniobra 13: ACONDICIONAMIENTO PARA BAJAR VARILLAS
            desc13 = "Acondicionar boca de pozo, montar piso de trabajo + herramientas de v/b."
            m.append({
                "MANIOBRA NORMALIZADA": "ACONDICIONAMIENTO PARA BAJAR VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "254",
                "Punto programa":       13,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc13
            })

            # Maniobra 14: BAJA VARILLAS
            desc14 = "Tareas varias durante bajada de varillas."
            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       14,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc14
            })

           # Maniobra 15: BAJA VARILLAS EN SIMPLE
            # Construyo “INSTALACIÓN FINAL varillas” a partir de varillas_fin en orden inverso
            parts = []
            for _, row in varillas_fin.iloc[::-1].iterrows():
                qty   = int(row["CANTIDAD"]) if pd.notna(row["CANTIDAD"]) else 0
                elem  = row["ELEMENTO"]
                cond  = f" {row['CONDICIÓN']}" if pd.notna(row["CONDICIÓN"]) else ""
                depth = f" en {int(row['PROFUNDIDAD'])}m" if pd.notna(row["PROFUNDIDAD"]) else ""
                diam  = f" de {row['DIÁMETRO']}"     if pd.notna(row["DIÁMETRO"])  else ""
                acero_vb   = f", {row['ACERO V/B']}"     if pd.notna(row["ACERO V/B"])     else ""
                cupla      = f", cupla {row['CUPLA SH/FS']}" if pd.notna(row["CUPLA SH/FS"]) else ""
                acero_cupla = f", {row['ACERO CUPLA']}"      if pd.notna(row["ACERO CUPLA"])   else ""
                comm = row["COMENTARIO"] if pd.notnull(row["COMENTARIO"]) else ""
                if "BAJA EN DOBLE" in comm.upper():
                    comm_tag = " (EN DOBLE)"
                elif "BAJA EN SIMPLE" in comm.upper():
                    comm_tag = " (BAJA EN SIMPLE)"
                else:
                    comm_tag = ""

                piece = f"{qty} {elem}{cond}{diam}{depth}{acero_vb}{cupla}{acero_cupla}{comm_tag}"
                parts.append(piece.strip())

            diseño_fin_vb = " + ".join(parts)

            desc15 = (
                "Tomar datos de bomba y bajarla + sarta de v/b en simple limpiando todas las conexiones "
                "con detergente biodegradable. Realizar control de torque cada 15 varillas según grado "
                "de acero de varillas. "
                f"Diseño a bajar: {diseño_fin_vb}."
            )

            m.append({
                "MANIOBRA NORMALIZADA": "BAJA VARILLAS EN SIMPLE",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP16",
                "ACTIVITY_SUBCODE":     "255",
                "Punto programa":       15,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc15
            })


            # 26) ACONDICIONA PARA PH
            m.append({
                "MANIOBRA NORMALIZADA":"ACONDICIONA PARA PH",
                "Punto programa":16,
                "DESCRIPCIÓN":"Acondicionar superficie para realizar prueba hidráulica.",
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"259","TIEMPO":""
            })


            # 27) PRUEBA DE HERMETICIDAD
            m.append({
                "MANIOBRA NORMALIZADA":"PRUEBA DE HERMETICIDAD",
                "Punto programa":17,
                "DESCRIPCIÓN":(
                    "Realizar PH final sobre bomba con 1000 psi. Junto a la prueba final, realizar prueba de funcionamiento de bomba. "
                    "Registrar la misma en OpenWells. Si la prueba es deficiente informar a Supervisor de Pulling."
                ),
                "ACTIVITY_PHASE":"SP03","ACTIVITY_CODE":"SP13","ACTIVITY_SUBCODE":"205","TIEMPO":""
            })



            # Maniobra 16: VARIOS
            desc16 = "Tareas varias."
            m.append({
                "MANIOBRA NORMALIZADA": "VARIOS",
                "ACTIVITY_PHASE":       "SPV",
                "ACTIVITY_CODE":        "SPV",
                "ACTIVITY_SUBCODE":     "SPV",
                "Punto programa":       16,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc16
            })

            # Maniobra 17: ARMA BDP
            desc17 = "Arma BDP."
            m.append({
                "MANIOBRA NORMALIZADA": "ARMA BDP",
                "ACTIVITY_PHASE":       "SP04",
                "ACTIVITY_CODE":        "SP15",
                "ACTIVITY_SUBCODE":     "260",
                "Punto programa":       17,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc17
            })

            # Maniobra 18: DESMONTA EQUIPO
            desc18 = (
                "Acondicionar boca de pozo, material sobrante y locación, instalar rotador de varillas y accesorios de superficie. "
                "Desmontar. Informar a Coordinación y Sala de Monitoreo de Pulling la finalización de la intervención y transporte a próxima locación. "
                "Generar acta de entrega/recepción de locación. Indicar si el puente de producción queda armado."
            )
            m.append({
                "MANIOBRA NORMALIZADA": "DESMONTA EQUIPO",
                "ACTIVITY_PHASE":       "SP01",
                "ACTIVITY_CODE":        "SP11",
                "ACTIVITY_SUBCODE":     "202",
                "Punto programa":       18,
                "TIEMPO":               "",
                "DESCRIPCIÓN":          desc18
            })

            # 2) Renumerar de 1 a 18 (por si hubiera quedado algún "Punto programa" fuera de orden)
            for i, row in enumerate(m, start=1):
                row["Punto programa"] = i

            # 3) Construir DataFrame a partir de m y mostrar + guardar en Excel
            df_cambio = pd.DataFrame(m)
            display(df_cambio)
            df_cambio.to_excel("cambio_bes_x_bm_program.xlsx", index=False)
            print("\n✅ Archivo guardado como 'cambio_bes_x_bm_program.xlsx'")
            
       

            program_dropdown.observe(on_program_change, names='value')

        btn2.layout.display = ''
        out2.layout.display = ''
        display(btn2, out2)   

btn1.on_click(on_btn1_click)        
      



import ipywidgets as widgets
from IPython.display import display, clear_output
from zipfile import BadZipFile


# 2) Al hacer clic, ejecuto TODO el bloque “oculto”
def on_btn2_click(b):
    with out2:
        clear_output()   # limpia salidas previas

        # ——— Aquí va todo tu código original de la Celda 3, pero SIN volver a redeclarar otro btn u out ———

        import os
        import glob
        import pandas as pd
        import numpy as np
        import joblib
        import xgboost as xgb
        from sklearn.preprocessing import FunctionTransformer
        from IPython.display import display  # ya importaste clear_output arriba
        import ipywidgets as widgets  # si terminas usando widgets dentro de esta sección

        # 0) Cargo coords_df para lat/lon por POZO
        coords_path = r"C:\Users\ry16123\OneDrive - YPF\Escritorio\power BI\GUADAL- POWER BI\Inteligencia Artificial\Carga Programas de Pu\prediccion de HS\coordenadas1.xlsx"
        coords_df = pd.read_excel(coords_path, dtype={"POZO": str})
        for c in ("GEO_LATITUDE", "GEO_LONGITUDE"):
            coords_df[c] = (
                coords_df[c]
                .astype(str)
                .str.replace(",", ".", regex=False)
                .astype(float)
            )

        # 1) Cargo modelos y mappings
        model1   = joblib.load(r"C:\Users\ry16123\OneDrive - YPF\Escritorio\power BI\GUADAL- POWER BI\Inteligencia Artificial\Carga Programas de Pu\DATA SHEET\model1.pkl")
        mapping1 = joblib.load(r"C:\Users\ry16123\OneDrive - YPF\Escritorio\power BI\GUADAL- POWER BI\Inteligencia Artificial\Carga Programas de Pu\DATA SHEET\mapping1.pkl")
        model2   = joblib.load(r"C:\Users\ry16123\OneDrive - YPF\Escritorio\power BI\GUADAL- POWER BI\Inteligencia Artificial\Carga Programas de Pu\DATA SHEET\model2.pkl")
        mapping2 = joblib.load(r"C:\Users\ry16123\OneDrive - YPF\Escritorio\power BI\GUADAL- POWER BI\Inteligencia Artificial\Carga Programas de Pu\DATA SHEET\mapping2.pkl")

        FEATURES1 = ["rig_name","Eventos Normalizados","MANIOBRAS NORMALIZADAS","GEO_LATITUDE","GEO_LONGITUDE"]
        FEATURES2 = ["rig_name","Eventos Normalizados","MANIOBRAS NORMALIZADAS","pickup_weight","GEO_LATITUDE","GEO_LONGITUDE"]

        log_tr = FunctionTransformer(np.log1p, inverse_func=np.expm1)

        def apply_fe_new(df_new, mapping):
            df = df_new.copy()
            for col, freq_map in mapping.items():
                if col in df.columns:
                    df[col] = df[col].map(freq_map).fillna(0)
            return df

        def predict_single(inputs: dict) -> float:
            mani = inputs["MANIOBRAS NORMALIZADAS"]
            if mani in mapping2.get("MANIOBRAS NORMALIZADAS", {}):
                mdl, mp, feats = model2, mapping2, FEATURES2
            else:
                mdl, mp, feats = model1, mapping1, FEATURES1
            df_tmp = pd.DataFrame([inputs])[feats]
            df_enc = apply_fe_new(df_tmp, mp)
            dmat = xgb.DMatrix(df_enc)
            pred_t = mdl.predict(dmat)
            return float(log_tr.inverse_transform(pred_t.reshape(-1,1))[0,0])

        def fill_times(df, rig_name, evento, pozo, pickup_weights):
            motivos_falla_bm = {
                "Recupero BM",
                "Instalación final BM",
                "Cambio vástago/ajuste de medida",
                "Pesca de varilla"
            }
            evento_norm = "FALLA BM" if evento in motivos_falla_bm else evento

            sel = coords_df.loc[coords_df["POZO"] == str(pozo), ["GEO_LATITUDE", "GEO_LONGITUDE"]]
            if sel.empty:
                raise KeyError(f"No encontré coordenadas para POZO={pozo!r}")
            lat, lon = sel.iloc[0]

            print("=== Inputs generales para la predicción ===")
            print(f"POZO             : {pozo}")
            print(f"Rig name         : {rig_name}")
            print(f"Evento original  : {evento!r}")
            print(f"Evento normalizado: {evento_norm!r}")
            print(f"Latitud          : {lat}")
            print(f"Longitud         : {lon}")
            print("==========================================\n")

            df_out = df.copy()
            tiempos = []
            for mani in df_out["MANIOBRAS NORMALIZADAS"]:
                inp = {
                    "rig_name":               rig_name,
                    "Eventos Normalizados":   evento_norm,
                    "MANIOBRAS NORMALIZADAS": mani,
                    "pickup_weight":          float(pickup_weights.get(mani, 0)),
                    "GEO_LATITUDE":           lat,
                    "GEO_LONGITUDE":          lon
                }
                t = predict_single(inp)
                # redondear al cuarto de hora
                t = round(t * 4) / 4
                tiempos.append(t)

            df_out["TIEMPO"] = [f"{t:.2f}".replace(".", ",") for t in tiempos]

            total = sum(tiempos)
            total_str = f"{total:.2f}".replace(".", ",")
            total_row = {col: "" for col in df_out.columns}
            total_row["MANIOBRAS NORMALIZADAS"] = "Total"
            total_row["TIEMPO"] = total_str
            df_out = pd.concat([df_out, pd.DataFrame([total_row])], ignore_index=True)

            return df_out

        # 2) Encuentro el último “*_program.xlsx” que generó la Celda 2
        candidatos = glob.glob("*_program.xlsx")
        if not candidatos:
            raise FileNotFoundError("No encuentro archivos '*_program.xlsx' – ¿corriste la Celda 2?")
        program_file = max(candidatos, key=os.path.getmtime)
        print(f"📄 Usando programa base: '{program_file}'")

        # 3) Cargo el programa y normalizo los nombres de columna
        df_prog = pd.read_excel(program_file)

        # 3.1) Limpiar espacios y pasar a mayúsculas todos los nombres de columna
        df_prog.columns = df_prog.columns.str.strip().str.upper()

        # 3.2) Asegurar que exista una columna llamada "MANIOBRAS NORMALIZADAS"
        if "MANIOBRAS NORMALIZADAS" not in df_prog.columns:
            if "MANIOBRA NORMALIZADA" in df_prog.columns:
                df_prog.rename(columns={"MANIOBRA NORMALIZADA": "MANIOBRAS NORMALIZADAS"}, inplace=True)
            else:
                raise KeyError("El programa no tiene ni 'MANIOBRAS NORMALIZADAS' ni 'MANIOBRA NORMALIZADA'")

        # 3.3) Asegurar que exista una columna llamada "PUNTO PROGRAMA"
        if "PUNTO PROGRAMA" not in df_prog.columns:
            # Buscamos cualquier columna cuyo nombre, tras quitar espacios, coincida con PUNTOPROGRAMA
            for c in df_prog.columns:
                if c.replace(" ", "") == "PUNTOPROGRAMA":
                    df_prog.rename(columns={c: "PUNTO PROGRAMA"}, inplace=True)
                    break
            else:
                # Si no existe, la creamos directamente
                df_prog["PUNTO PROGRAMA"] = np.arange(1, len(df_prog) + 1)
        else:
            # Si ya existe, la sobreescribimos con valores 1..n
            df_prog["PUNTO PROGRAMA"] = np.arange(1, len(df_prog) + 1)

        print("➤ Vista previa (sin tiempos):")
        display(df_prog.head())

        
        def default_weight_for_maniobra(mani):
            suma_varillas_act = varillas_act["CANTIDAD"].fillna(0).astype(float).sum()
            suma_varillas_fin = varillas_fin["CANTIDAD"].fillna(0).astype(float).sum()
            suma_tubing_act   = tubing_act["CANTIDAD"].fillna(0).astype(float).sum()
            suma_tubing_fin   = tubing_fin["CANTIDAD"].fillna(0).astype(float).sum()

            if mani == "SACA VARILLAS EN PESCA":
                return suma_varillas_act / 2.0
            elif mani in {
                "SACA VARILLAS EN PESCA EN DOBLE",
                "SACA VARILLAS EN PESCA EN SIMPLE",
                "SACA VARILLAS EN DOBLE",
                "SACA VARILLAS EN SIMPLE"
            }:
                return suma_varillas_act

            elif mani == "SACA TUBING EN PESCA":
                return suma_tubing_act / 2.0
            elif mani in {
                "SACA TUBING DESAGOTANDO",
                "SACA TUBING BES",
                "SACA TUBING EN DOBLE",
                "SACA TUBING EN SIMPLE",
                "SACA TUBING CALIBRANDO EN DOBLE",
                "SACA TUBING CALIBRANDO EN SIMPLE"
            }:
                return suma_tubing_act

            elif mani == "BAJA TUBING DESAGOTANDO":
                return suma_tubing_fin
            elif mani in {
                "BAJA TUBING BES",
                "BAJA TUBING EN DOBLE",
                "BAJA TUBING EN SIMPLE",
                "BAJA CALIBRANDO TUBING EN DOBLE",
                "BAJA CALIBRANDO TUBING EN SIMPLE"
            }:
                return suma_tubing_fin

            elif mani in {
                "BAJA VARILLAS EN PESCA EN DOBLE",
                "BAJA VARILLAS EN PESCA EN SIMPLE",
                "BAJA VARILLAS EN DOBLE",
                "BAJA VARILLAS EN SIMPLE"
            }:
                return suma_varillas_fin

            return 0.0

        # ————————————————————————————————
        # 4) Formulario para ingresar pickup_weights SOLO en las maniobras indicadas:
        allowed_pickups = {
            "SACA VARILLAS EN PESCA",
            "SACA VARILLAS EN PESCA EN DOBLE",
            "SACA VARILLAS EN PESCA EN SIMPLE",
            "SACA VARILLAS EN DOBLE",
            "SACA VARILLAS EN SIMPLE",
            "SACA TUBING DESAGOTANDO",
            "SACA TUBING EN PESCA",
            "SACA TUBING BES",
            "SACA TUBING EN DOBLE",
            "SACA TUBING EN SIMPLE",
            "SACA TUBING CALIBRANDO EN DOBLE",
            "SACA TUBING CALIBRANDO EN SIMPLE",
            "BAJA TUBING DESAGOTANDO",
            "BAJA TUBING BES",
            "BAJA TUBING EN DOBLE",
            "BAJA TUBING EN SIMPLE",
            "BAJA CALIBRANDO TUBING EN DOBLE",
            "BAJA CALIBRANDO TUBING EN SIMPLE",
            "BAJA VARILLAS EN PESCA EN DOBLE",
            "BAJA VARILLAS EN PESCA EN SIMPLE",
            "BAJA VARILLAS EN DOBLE",
            "BAJA VARILLAS EN SIMPLE",
        }

        all_manis = df_prog["MANIOBRAS NORMALIZADAS"].unique().tolist()
        mani_para_peso = [m for m in all_manis if m in allowed_pickups]

        weight_inputs = {}
        for m in mani_para_peso:
            default_val = default_weight_for_maniobra(m)
            weight_inputs[m] = widgets.FloatText(
                value=default_val,
                description=m,
                step=0.1,
                style={"description_width": "initial"}
            )

        print("👉 Ingresa (o corrige) los pesos (lbs) para cada maniobra que aplique:")
        if weight_inputs:
            display(widgets.VBox(list(weight_inputs.values())))
        else:
            print("No hay ninguna maniobra que requiera pickup_weight en este programa.")

        # 5) Botón y salida para “Predecir TIEMPOS” (reemplaza el widget que estaba en la Celda 3)
        btn_pred = widgets.Button(description="Predecir TIEMPOS", button_style="success")
        out_pred = widgets.Output()

        def _on_click_pred(b):
            with out_pred:
                clear_output()
                pickup_weights = {m: w.value for m, w in weight_inputs.items() if w.value > 0}
                df_with_times = fill_times(
                    df_prog,
                    rig_name=meta["EQUIPO"],
                    evento=meta["MANIOBRAS (MOTIVO)"],
                    pozo=meta["POZO"],
                    pickup_weights=pickup_weights
                )

                # ————— Renombrar columnas a los nombres finales —————
                df_with_times = df_with_times.rename(
                    columns={"MANIOBRAS NORMALIZADAS": "MANIOBRA NORMALIZADA"}
                )

                # 1) Reordenar/filtrar columnas antes de procesar "VARIOS"
                cols_orden = [
                    "MANIOBRA NORMALIZADA",
                    "ACTIVITY_PHASE",
                    "ACTIVITY_CODE",
                    "ACTIVITY_SUBCODE",
                    "PUNTO PROGRAMA",
                    "TIEMPO",
                    "DESCRIPCIÓN"
                ]
                cols_existentes = [c for c in cols_orden if c in df_with_times.columns]
                df_with_times = df_with_times[cols_existentes]

                
                               
                # 2) Trabajar sobre la columna TIEMPO:
                #     – Extraer fila "VARIOS"
                #     – Dividir su TIEMPO entre todas las filas restantes (excluyendo "Total")
                #     – Sumar esa parte a cada fila “real” y eliminar “VARIOS”
                #     – Recalcular el total y actualizar la fila “Total”
                df_mod = df_with_times.copy()
                
                
                # (a) Separar la fila "Total" si existe
                mask_total = df_mod["MANIOBRA NORMALIZADA"] == "Total"
                if mask_total.any():
                    total_row = df_mod.loc[mask_total].copy()
                    df_mod = df_mod.loc[~mask_total]
                else:
                    total_row = None

                # (b) Extraer la fila "VARIOS" (si existe)
                mask_varios = df_mod["MANIOBRA NORMALIZADA"] == "VARIOS"
                if mask_varios.any():
                    # Convertir el TIEMPO de "VARIOS" a float (coma → punto)
                    varios_str = df_mod.loc[mask_varios, "TIEMPO"].iloc[0]
                    varios_time = float(varios_str.replace(",", "."))

                    # Eliminar la fila “VARIOS”
                    df_mod = df_mod.loc[~mask_varios]

                    # Cantidad de filas restantes (solo maniobras reales)
                    n = len(df_mod)

                    # Incremento a sumar en cada tiempo = varios_time / n
                    incremento = (varios_time / n) if n > 0 else 0.0

                    # Función para redondear al cuarto de hora
                    def a_cuarto_de_hora(x: float) -> float:
                        return round(x * 4) / 4

                    # (c) Convertir cada TIEMPO de cadena a float, sumarle el incremento y redondear
                    tiempos_flotantes = (
                        df_mod["TIEMPO"]
                        .str.replace(",", ".")
                        .astype(float)
                        .add(incremento)
                        .apply(a_cuarto_de_hora)
                    )

                    # (d) Volver a formatear como cadena con coma decimal
                    df_mod["TIEMPO"] = tiempos_flotantes.apply(lambda t: f"{t:.2f}".replace(".", ","))
                
                 # — 1) Forzar valores por defecto en TIEMPO para dos filas específicas —
                # — 1) Forzar valores por defecto en TIEMPO para dos filas específicas —
                mask_equipo  = df_mod["MANIOBRA NORMALIZADA"] == "EQUIPO EN TRANSPORTE"
                mask_montaje = df_mod["MANIOBRA NORMALIZADA"] == "MONTAJE EQUIPO"

                # Obtenemos el tipo de equipo desde el datasheet
                equip_type = meta.get("EQUIPO", "").upper()

                if equip_type == "PU":
                    df_mod.loc[mask_equipo,  "TIEMPO"] = "3,50"
                    df_mod.loc[mask_montaje, "TIEMPO"] = "3,50"
                elif equip_type == "FB":
                    df_mod.loc[mask_equipo,  "TIEMPO"] = "2,00"
                    df_mod.loc[mask_montaje, "TIEMPO"] = "1,50"
                
                
                # — 2) Ajuste de tiempos de PH y Hermeticidad: sólo en la ÚLTIMA aparición —
                if equip_type == "PU":
                    # última "ACONDICIONA PARA PH"
                    mask_ph = df_mod["MANIOBRA NORMALIZADA"] == "ACONDICIONA PARA PH"
                    if mask_ph.any():
                        idx_ph = df_mod.index[mask_ph].max()
                        df_mod.at[idx_ph, "TIEMPO"] = "0,75"

                    # última "PRUEBA DE HERMETICIDAD"
                    mask_test = df_mod["MANIOBRA NORMALIZADA"] == "PRUEBA DE HERMETICIDAD"
                    if mask_test.any():
                        idx_test = df_mod.index[mask_test].max()
                        df_mod.at[idx_test, "TIEMPO"] = "1,25"
                
                # Detectar si en la instalación final de tubing hay “SHEAR OUT”
                cond_shear = (
                    tubing_fin["ELEMENTO"].str.contains("SHEAR", case=False, na=False).any()
                    or tubing_fin["COMENTARIO"].str.contains("SHEAR", case=False, na=False).any()
                )

                # —————— Ajuste para el RESTO de las apariciones de “ACONDICIONA PARA PH” ——————
                mask_ph_all  = df_mod["MANIOBRA NORMALIZADA"] == "ACONDICIONA PARA PH"
                idx_last_ph  = df_mod.index[mask_ph_all].max()  # índice de la última aparición
                mask_ph_rest = mask_ph_all & (df_mod.index != idx_last_ph)

                if cond_shear:
                    df_mod.loc[mask_ph_rest, "TIEMPO"] = "0,50"
                else:
                    df_mod.loc[mask_ph_rest, "TIEMPO"] = "1,25"

                # —————— Ajuste para el RESTO de las apariciones de “PRUEBA DE HERMETICIDAD” ——————
                mask_ht_all  = df_mod["MANIOBRA NORMALIZADA"] == "PRUEBA DE HERMETICIDAD"
                idx_last_ht  = df_mod.index[mask_ht_all].max()  # índice de la última aparición
                mask_ht_rest = mask_ht_all & (df_mod.index != idx_last_ht)

                if cond_shear:
                    df_mod.loc[mask_ht_rest, "TIEMPO"] = "0,75"
                else:
                    df_mod.loc[mask_ht_rest, "TIEMPO"] = "2,00"        
                
                

                # — 3) Recalcular la fila “Total” con los nuevos tiempos —
                #    (el resto de tu código de totalización queda igual)
                tiempos_sum = (
                    df_mod["TIEMPO"]
                    .str.replace(",", ".")
                    .astype(float)
                    .sum()
                )
                total_redondeado = round(tiempos_sum * 4) / 4
                total_str       = f"{total_redondeado:.2f}".replace(".", ",")

                total_row_nueva = {col: "" for col in df_mod.columns}
                total_row_nueva["MANIOBRA NORMALIZADA"] = "Total"
                total_row_nueva["TIEMPO"]               = total_str

                total_row_df = pd.DataFrame([total_row_nueva], columns=df_mod.columns)
                df_final     = pd.concat([df_mod, total_row_df], ignore_index=True)

                print("✅ Tiempos predichos (ajustados y redondeados):")
                display(df_final)

                # ———————————— NUEVAS LÓGICAS ANTES DE GUARDAR ————————————
                
                #logica de altura de vastago. 
                # ————— LÓGICA PARA DEJAR ALTURA DE VÁSTAGO EN “DESMONTA EQUIPO” —————
                # Buscamos si en la tabla Final de Varillas hay VASTAGO con profundidad > 0
                mask_vastago = (
                    varillas_fin["ELEMENTO"]
                    .str.contains("VASTAGO", case=False, na=False)
                )
                if mask_vastago.any():
                    # Tomamos la primera profundidad no-nula
                    profund = (
                        varillas_fin.loc[mask_vastago, "PROFUNDIDAD"]
                        .dropna()
                        .astype(float)
                        .abs()
                    )
                    if not profund.empty:
                        altura = profund.iloc[0]
                        # Localizamos la fila “DESMONTA EQUIPO” en el df_mod
                        mask_desmonta = df_mod["MANIOBRA NORMALIZADA"] == "DESMONTA EQUIPO"
                        if mask_desmonta.any():
                            idx = df_mod.index[mask_desmonta][0]
                            texto_adicional = (
                                f"Dejar altura de vástago. {altura} mts "
                                "(medido de nivel de terreno, hasta nivel de cupla de vástago). "
                            )
                            # Prepend o append, según prefieras
                            df_mod.at[idx, "DESCRIPCIÓN"] = (
                                texto_adicional + df_mod.at[idx, "DESCRIPCIÓN"]
                            )


                
                
                
                # 1) Reemplazo "Desclavar bomba" según INSTALACIÓN ACTUAL DE TUBING
                if tubing_act["ELEMENTO"].str.contains("CAMISA TH", case=False, na=False).any():
                    df_mod["DESCRIPCIÓN"] = df_mod["DESCRIPCIÓN"].str.replace(
                        "Desclavar bomba",
                        "Librar Pistón TH",
                        regex=False
                    )
                elif tubing_act["ELEMENTO"].str.contains("ON-OFF", case=False, na=False).any():
                    df_mod["DESCRIPCIÓN"] = df_mod["DESCRIPCIÓN"].str.replace(
                        "Desclavar bomba",
                        "Librar ON-OFF",
                        regex=False
                    )

                # 2) Lógicas para "Tomar datos de bomba y bajarla"
                tubing_fin["ELEMENTO"]   = tubing_fin["ELEMENTO"].str.strip()
                varillas_fin["ELEMENTO"] = varillas_fin["ELEMENTO"].str.strip()

                cond_tub_fin    = tubing_fin["ELEMENTO"].str.contains("CAMISA TH", case=False, na=False).any()
                cond_var_piston = varillas_fin["ELEMENTO"].str.contains("PISTON",   case=False, na=False).any()
                cond_var_onoff  = varillas_fin["ELEMENTO"].str.contains("ON-OFF",   case=False, na=False).any()

                if cond_tub_fin and cond_var_piston:
                    df_mod["DESCRIPCIÓN"] = df_mod["DESCRIPCIÓN"].str.replace(
                        "Tomar datos de bomba y bajarla",
                        "Bajar PISTÓN DE TH ",
                        regex=False
                    )
                elif cond_var_onoff:
                    df_mod["DESCRIPCIÓN"] = df_mod["DESCRIPCIÓN"].str.replace(
                        "Tomar datos de bomba y bajarla",
                        "Bajar conector ON-OFF ",
                        regex=False
                    )

                # 3) Si hay SHEAR + BAJA en instalación final de tubing, ajustamos la descripción de PH
                cond_shear = tubing_fin["ELEMENTO"].str.contains("SHEAR", case=False, na=False).any()
                cond_baja  = tubing_fin["COMENTARIO"].str.contains("BAJA", case=False, na=False).any()

                if cond_shear and cond_baja:
                    df_mod["DESCRIPCIÓN"] = df_mod["DESCRIPCIÓN"].str.replace(
                        "Realizar PH inicial, intermedia y final con 1000, 900 y 800 psi",
                        "Realizar PH inicial, intermedia y final con 1000, 900 y 800 psi. Liberar shear out con 2000 psi",
                        regex=False
                    )

                # ————— LÓGICA PARA “FIJA ANCLA” (antes de recalcular total) —————
                # ————— LÓGICA PARA “FIJA ANCLA” (antes de recalcular total) —————
                nombre_prog = os.path.basename(program_file)
                # Solo ejecutamos la eliminación si NO es el programa de cambio BBA-TUBING con ancla
                if nombre_prog != "cambio_bba_tubing_con_ancla_program.xlsx":
                    # Si en la instalación final NO hay ANCLA, quitamos la fila FIJA ANCLA
                    if not tubing_fin["ELEMENTO"].astype(str).str.contains("ANCLA", case=False, na=False).any():
                        df_mod = df_mod[df_mod["MANIOBRA NORMALIZADA"] != "FIJA ANCLA"].reset_index(drop=True)

                # ————— LÓGICA “Profundizar tubing” —————
                reqs = [meta.get(f"REQ_ESP_{i}", "") for i in range(1, 5)]
                if any(isinstance(r, str) and "Profundizar" in r for r in reqs):
                    # ¿Hay ancla en la instalación actual?
                    tiene_ancla_act = tubing_act["ELEMENTO"].astype(str).str.contains("ANCLA", case=False, na=False).any()
                    # localizamos la primera ACONDICIONAMIENTO PARA SACAR CAÑOS
                    mask_acond = df_mod["MANIOBRA NORMALIZADA"] == "ACONDICIONAMIENTO PARA SACAR CAÑOS"
                    if mask_acond.any():
                        idx = df_mod.index[mask_acond][0]
                        # construimos la descripción con o sin prefijo
                        prefijo = "Librar ancla. " if tiene_ancla_act else ""
                        descripcion = prefijo + "Profundizar hasta contactar fondo- Informar resultados para ver como continuar."
                        nueva_fila = {
                            "MANIOBRA NORMALIZADA": "BAJA TUBING",
                            "ACTIVITY_PHASE":       "SP04",
                            "ACTIVITY_CODE":        "SP16",
                            "ACTIVITY_SUBCODE":     "257",
                            "PUNTO PROGRAMA":       "",       # se recalculará luego
                            "TIEMPO":               "1,2",
                            "DESCRIPCIÓN":          descripcion
                        }
                        # completamos resto de columnas con ""
                        for c in df_mod.columns:
                            if c not in nueva_fila:
                                nueva_fila[c] = ""
                        # insertamos justo después de idx
                        df_mod = pd.concat([
                            df_mod.iloc[:idx+1],
                            pd.DataFrame([nueva_fila], columns=df_mod.columns),
                            df_mod.iloc[idx+1:]
                        ], ignore_index=True)

                        # ——— Limpiar cualquier “Librar ancla.” redundante en filas posteriores ———
                        mask_resto = df_mod.index > idx+1
                        df_mod.loc[mask_resto, "DESCRIPCIÓN"] = (
                            df_mod.loc[mask_resto, "DESCRIPCIÓN"]
                                 .str.replace(r'(?i)\blibrar ancla\.\s*', '', regex=True)
                        )
                # ————— Capar tiempos máximos de ciertas maniobras y avisar —————
                cap_limits = {
                    "ACONDICIONA PARA SACAR VARILLAS":        3.5,
                    "DESARMA BDP":                            1.5,
                    "CIRCULA":                                2.25,
                    "ACONDICIONAMIENTO PARA SACAR CAÑOS":      4.5,
                    "ACONDICIONAMIENTO PARA BAJAR VARILLAS":  1.5,
                    "ARMA BDP":                               1.5,
                }

                for maniobra, limite in cap_limits.items():
                    # Máscara de filas afectadas
                    mask = df_mod["MANIOBRA NORMALIZADA"] == maniobra
                    if not mask.any():
                        continue

                    # Parseamos tiempos a float
                    tiempos = (
                        df_mod.loc[mask, "TIEMPO"]
                        .str.replace(",", ".")
                        .astype(float)
                    )
                    # Detectamos cuáles superan el límite
                    over = tiempos > limite
                    if over.any():
                        # Por cada fila que excede, imprimimos un aviso
                        for idx, orig in tiempos[over].items():
                            print(f"⚠ Se capó '{maniobra}' de {orig:.2f} h a {limite:.2f} h (fila índice {idx})")
                        # Clippeamos y formateamos de vuelta
                        nuevos = tiempos.clip(upper=limite).map(lambda t: f"{t:.2f}".replace(".", ","))
                        df_mod.loc[mask, "TIEMPO"] = nuevos
                
                        
                        
                # ————— REASIGNO PUNTO PROGRAMA Y RECALCULO TOTAL —————
               # REASIGNO PUNTO PROGRAMA
                df_mod["PUNTO PROGRAMA"] = np.arange(1, len(df_mod) + 1)

                # RECALCULO TOTAL
                tiempos   = df_mod["TIEMPO"].str.replace(",", ".").astype(float).sum()
                total_red = round(tiempos * 4) / 4
                total_str = f"{total_red:.2f}".replace(".", ",")

                total_row = {c: "" for c in df_mod.columns}
                total_row["MANIOBRA NORMALIZADA"] = "Total"
                total_row["TIEMPO"]               = total_str

                df_final = pd.concat([df_mod, pd.DataFrame([total_row])], ignore_index=True)

                # … finalmente guardas df_final …
                out_file = program_file.replace(".xlsx", "_con_tiempos.xlsx")
                df_final.to_excel(out_file, index=False)
                print(f"\n📦 Guardado: '{out_file}'")
                
                # ————————————— FIN DE LAS NUEVAS LÓGICAS ——————————————
        btn_pred.on_click(_on_click_pred)
        display(btn_pred, out_pred)

                
                
btn2.on_click(on_btn2_click)    

                

# 3) Asocio el botón principal al callback y lo muestro

display(btn1, out1)


# In[24]:


#------------------CODIGO PARA EXPORTAR LA INSTALACION FINAL A OW-----------------------------------------------------

import pandas as pd
from openpyxl import load_workbook


# ---------------------- FUNCIONES AUXILIARES ----------------------
def transformar_tipo_seccion(valor):
    if pd.isna(valor): return ''
    v = str(valor).upper()
    if 'TUBING - ACERO' in v or 'RED.- 3,500" X 2,875"' in v or 'RED.- 2,875" X 3.500"' in v or 'RED.- 2,875" X 2,875"' in v:
        return 'TUBING'
    if 'BBA. PCP' in v: return 'BOMBA CAVIDAD PROGRESIVA'
    if 'BBA. MECANICA' in v or 'BBA. TUB.PUMP' in v:
        return 'BOMBA CONVENCIONAL'
    if 'BBA. ELECT.' in v: return 'BOMBA ELECTROSUMERGIBLE'
    if 'PACKER' in v: return 'PACKER'
    if 'TAPON' in v or 'SHEAR - OUT' in v or 'ZAPATO A COPAS' in v or 'ANCLA TUBING' in v or 'ANCLA CSG' in v:
        return 'EQUIPAMIENTO DE POZO'

    return valor

def transformar_tipo_componente(valor):
    if pd.isna(valor): return ''
    v = str(valor).upper()
    m = {
        'SHEAR - OUT':'SHEAR OUT',
        'ZAPATO A COPAS':'ZAPATO',
        'RED.- 3,500" X 2,875"':'REDUCCION',
        'RED.- 2,875" X 2,875"':'REDUCCIÓN',
        'RED.- 2,875" X 3.500"':'REDUCCIÓN',
        'TUBING - ACERO':'TUBING',
        'BBA. PCP - ESTATOR':'ESTATOR',
        'BBA. PCP - NIPLE DE PARO':'NIPLE DE PARO',
        'ANCLA TUBING - DE TORQUE':'ANCLA',
        'ANCLA TUBING':'ANCLA',
        'BBA. ELECT.- NIPLE DE PURGA':'NIPLE DE PURGA',
        'BBA. ELECT.- VALV.RETENSION':'VALVULA DE RETENCION',
        'BBA. ELECT.- CAB.DESCARGA':'CABEZA DESCARGA',
        'BBA. ELECT.- BOMBA':'BOMBA',
        'BBA. ELECT.- ENTRADA':'INTAKE',
        'BBA. ELECT.- SELLO':'SELLO',
        'BBA. ELECT.- MOTOR':'MOTOR',
        'BBA. ELECT.- SENSOR P.S.I.':'SENSOR DE PRESION',
        'BBA. TUB.PUMP':'TUBING PUMP'
    }
    for k, v2 in m.items():
        if k in v: return v2
    return valor

# ---------------------- CREACIÓN PESTAÑA OW -----------------------------------------------------------------------------------------------
def crear_hoja_ow(libro, hoja_ds):
    # 1) Determinar "Nombre de la columna"
    vals = [hoja_ds[f'K{i}'].value for i in range(59, 77)]
    vals = [str(v).upper() for v in vals if isinstance(v, str)]
    nombre_col = ''
    for v in vals:
        if 'BBA. PCP' in v:
            nombre_col = 'SARTA VARILLAS -PCP-'
            break
        if 'BBA. TUB.PUMP' in v or 'BBA. MECANICA' in v:
            nombre_col = 'SARTA VARILLAS -BM-'
            break

    # 2) DataFrame base
    df = pd.DataFrame({
        'SideTrack N°':['OH'],
        'Nombre de la columna':[nombre_col],
        'Diámetro Nominal (in)':[''],
        'MD Desde (m)':[''],
        'Corre dentro de:':[''],
        'Longitud (m)':[''],
        'MD Total (m)':[''],
        'Mínimo I.D. (in)':[''],
        'Descripción':['INSTAL.DE PRODUCCION'],
        'Instalación Dual?':[''],
        'Comentarios':['']
    })

    # 3) Crear/reemplazar hoja
    if 'OW' in libro.sheetnames: del libro['OW']
    ws = libro.create_sheet('OW')
    for col, header in enumerate(df.columns, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=2, column=col, value=df.at[0, header])

    # 4) Valor "VARILLA BOMBEO - ACERO"
    val_acero = None
    for i in range(59, 78):
        k = hoja_ds[f'K{i}'].value
        if isinstance(k, str) and 'VARILLA BOMBEO - ACERO' in k.upper():
            m = hoja_ds[f'M{i}'].value
            try:
                val_acero = float(m)
            except: pass
            break
    if val_acero is not None:
        ws.cell(row=2, column=3, value=val_acero)

    # 5) Mayor "TUBING - ACERO"
    tubos = []
    for i in range(35, 55):
        k = hoja_ds[f'K{i}'].value
        if isinstance(k, str) and 'TUBING - ACERO' in k.upper():
            try: tubos.append(float(hoja_ds[f'M{i}'].value))
            except: pass
    if tubos and ws.cell(row=3, column=3).value is None:
        ws.cell(row=3, column=3, value=max(tubos))

    # 6) A3/I3
    if ws.cell(3,1).value is None: ws.cell(3,1,'OH')
    if ws.cell(3,9).value is None: ws.cell(3,9,'INSTAL.DE PRODUCCION')

    # 7) B3
    if ws.cell(3,2).value is None:
        for i in range(35,78):
            k = hoja_ds[f'K{i}'].value
            if isinstance(k,str):
                u = k.upper()
                if 'BBA. MECANICA' in u:
                    ws.cell(3,2,'SARTA TUBING -BM-'); break
                if 'BBA. PCP' in u:
                    ws.cell(3,2,'SARTA TUBING -PCP-'); break
                if 'BBA. ELECT.' in u:
                    ws.cell(3,2,'SARTA TUBING -ES-'); break

# ---------------------- CREACIÓN PESTAÑA OW1 ----------------------------------------------------------------------------------------------
def crear_hoja_ow1(libro):
    # Leer datos con pandas
    use = dict(sheet_name='Data Sheet', skiprows=32, nrows=21, header=None)
    df_k = pd.read_excel(ruta, usecols='K',  **use, names=['Componente'])
    df_l = pd.read_excel(ruta, usecols='L',  **use, names=['Condición'])
    df_m = pd.read_excel(ruta, usecols='M',  **use, names=['Diámetro Nominal (in)'])
    df_o = pd.read_excel(ruta, usecols='O',  **use, names=['Cantidad de piezas'])

    # Transformaciones
    df_k['Tipo Sección']    = df_k['Componente'].apply(transformar_tipo_seccion)
    df_k['Tipo Componente'] = df_k['Componente'].apply(transformar_tipo_componente)
    df_k['Condición']       = df_l['Condición']
    df_k['Diámetro Nominal (in)'] = df_m['Diámetro Nominal (in)']
    df_k['Cantidad de piezas']= df_o['Cantidad de piezas']

    # Cálculo de longitudes
    longitudes = []
    for comp, cant in zip(df_k['Componente'], df_k['Cantidad de piezas']):
        if pd.isna(comp) or pd.isna(cant):
            longitudes.append('')
            continue
        c = comp.upper()
        if 'TUBING - ACERO' in c: longitudes.append(cant * 9.3)
        elif 'ZAPATO' in c:       longitudes.append(cant * 0.4)
        elif 'RED.' in c:         longitudes.append(cant * 0.23)
        elif 'ANCLA' in c:        longitudes.append(cant * 0.86)
        elif 'SHEAR - OUT' in c:  longitudes.append(cant * 0.17)
        elif 'TAPON' in c:        longitudes.append(cant * 0.17)
        else:                     longitudes.append('')
    df_k['Longitud (m)'] = longitudes

    # Crear/reemplazar OW1
    if 'OW1' in libro.sheetnames: del libro['OW1']
    ws = libro.create_sheet('OW1')
    headers = [
        'Clip','Tipo Sección','Tipo Componente','Componente','Condición','Grado',
        'Diámetro Nominal (in)','Diám.int.(in)','Cantidad de piezas','Fabricante',
        'Longitud (m)','MD Tope / Desde (m)','MD Base / Hasta (m)','Calibre (in)',
        'Descripción Componente','Tipo de Material','Detalle','Tensión de Fijado (lbf)'
    ]
    for col, h in enumerate(headers,1):
        ws.cell(1,col,h)
    for i, row in df_k.iterrows():
        ws.cell(i+2,2, row['Tipo Sección'])
        ws.cell(i+2,3, row['Tipo Componente'])
        ws.cell(i+2,4, row['Componente'])
        ws.cell(i+2,5, row['Condición'])
        ws.cell(i+2,7, row['Diámetro Nominal (in)'])
        ws.cell(i+2,9, row['Cantidad de piezas'])
        ws.cell(i+2,11,row['Longitud (m)'])

# ---------------------- CREACIÓN PESTAÑA OW3 -----------------------------------------------------------------------------------------

def crear_hoja_ow3(libro, hoja_ds):
    eq_sec = {
        "VASTAGO":"VARILLA","TROZO VARILLA BOMBEO":"VARILLA",
        "VARILLA BOMBEO":"VARILLA","VARILLA DE PESO":"VARILLA","BARRA DE PESO":"VARILLA",
        "BBA. MECANICA":"BOMBA CONVENCIONAL","BBA. PCP":"BOMBA CAVIDAD PROGRESIVA", "BBA. TUB.PUMP":"BOMBA CONVENCIONAL"
    }
    eq_comp = {
        "VASTAGO":"VASTAGO","TROZO VARILLA BOMBEO":"TROZO VARILLA",
        "VARILLA BOMBEO - ACERO":"VARILLA DE BOMBEO","BARRA DE PESO":"VARILLA DE PESO", "VARILLA BOMBEO C/CENT.- ACERO":"VARILLA DE BOMBEO", "RED.- 2,875 X 3,500":"REDUCCION",
        "BBA. MECANICA":"CONVENCIONAL","VARILLA DE PESO":"VARILLA DE PESO",
        "BBA. TUB.PUMP":"TUBING PUMP","BBA. PCP - ROTOR":"ROTOR"
    }

    if 'OW3' in libro.sheetnames:
        hoja_dest = libro['OW3']
        hoja_dest.delete_rows(1, hoja_dest.max_row)
    else:
        hoja_dest = libro.create_sheet('OW3')
        headers = [
            "Clip","Tipo Sección","Tipo Componente","Componente","Condición","Grado",
            "Diámetro Nominal (in)","Diám.int.(in)","Cantidad de piezas","Fabricante",
            "Longitud (m)",'MD Tope / Desde (m)','MD Base / Hasta (m)','Calibre (in)',
            'Descripción Componente','Tipo de Material','Detalle','Tensión de Fijado (lbf)'
        ]
        hoja_dest.append(headers)

    for fila in range(57, 76):
        k = hoja_ds[f'K{fila}'].value
        if not k:
            continue

        l = hoja_ds[f'L{fila}'].value
        m = hoja_ds[f'M{fila}'].value
        r = hoja_ds[f'O{fila}'].value
        cantidad = hoja_ds[f'R{fila}'].value
        cantidad = cantidad if isinstance(cantidad, (int, float)) else 0

        K = str(k).upper()

        sec = next((v for k2, v in eq_sec.items() if k2 in K), '')
        comp = next((v for k2, v in eq_comp.items() if k2 in K), '')

        # Calcular longitud
        lon = ""
        if "VASTAGO - CROMADO" in K:
            lon = 6.7
        elif "VASTAGO - PCP" in K:
            lon = 3.04
        elif "BBA. MECANICA" in K:
            lon = 7.6
        elif "VARILLA BOMBEO" in K or "VARILLA DE PESO" in K:
            lon = round(cantidad * 7.62, 2)

        hoja_dest.append(["", sec, comp, k, l, r, m, "", cantidad, "", lon, "", "", "", "", "", ""])


# ---------------------- EJECUCIÓN PRINCIPAL ----------------------
def main():
    libro = load_workbook(ruta, keep_vba=True)
    ds = libro['Data Sheet']
    crear_hoja_ow(libro, ds)
    crear_hoja_ow1(libro)
    crear_hoja_ow3(libro, ds)
    libro.save(ruta)
    print("✅ Se crearon/actualizaron las hojas OW, OW1 y OW3 exitosamente.")

if __name__ == "__main__":
    main()


# In[ ]:




